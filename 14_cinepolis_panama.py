# Importamos el módulo 'os', que nos permite interactuar con el sistema operativo, como manejar archivos y directorios.
import os

# Importamos el módulo 'time', que nos permite trabajar con el tiempo, por ejemplo, hacer pausas en la ejecución del programa.
import time

# Importamos el módulo 'sys', que nos permite interactuar con el sistema de entrada y salida del programa, como leer argumentos de la línea de comandos.
import sys

# Importamos la clase 'datetime' del módulo 'datetime', que nos permite trabajar con fechas y horas.
from datetime import datetime

# Importamos el módulo 'openpyxl', que nos permite leer y escribir archivos de Excel.
import openpyxl

# Importamos las clases y funciones necesarias de 'selenium', que nos permiten automatizar la interacción con navegadores web.
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.edge.service import Service
from selenium.webdriver.edge.options import Options

# Importamos la función 'urlparse' del módulo 'urllib.parse', que nos permite analizar y descomponer URLs (direcciones web).
from urllib.parse import urlparse


# Creamos un objeto para configurar las opciones del navegador Microsoft Edge.
opciones_edge = Options()

# Configuramos el navegador para usar el motor de Chromium (que es el motor de Google Chrome).
opciones_edge.use_chromium = True

# Indicamos la ubicación del archivo del controlador de Microsoft Edge.
# Este archivo permite que el programa controle el navegador Edge.
servicio_edge = Service('C:\\Users\\William Monroy\\Desktop\\CODE\\msedgedriver.exe')

# Creamos un objeto que representa el navegador Microsoft Edge.
# Le decimos al programa dónde encontrar el controlador y qué opciones usar.
driver = webdriver.Edge(service=servicio_edge, options=opciones_edge)


# Creamos un nuevo libro de Excel en blanco.
libro_excel = openpyxl.Workbook()

# Verificamos si hay una hoja con el nombre 'Sheet' en el libro de Excel.
if 'Sheet' in libro_excel.sheetnames:
    # Si existe una hoja llamada 'Sheet', la eliminamos del libro de Excel.
    libro_excel.remove(libro_excel['Sheet'])
    


# La función `mostrar_carga` muestra una animación de carga en la consola.
# Se actualiza del 0% al 100% durante el tiempo especificado en segundos.
# La animación se muestra como un porcentaje que incrementa gradualmente.
# La función hace una pausa corta entre cada actualización para que el usuario
# pueda ver el progreso. Al final, muestra un mensaje indicando que la carga
# ha completado al 100%.
                        
def carga_completada(tiempo_total):
    
#1. Define el número total de pasos para la carga (100 pasos representan el 100% de la carga).
    numero_de_pasos = 100
    
#2. Calculamos el tiempo a esperar entre cada paso del indicador de carga.
    #tiempo_total es el tiempo total durante el cual debe mostrarse el indicador de carga.
    #numero_de_pasos es el número total de pasos (en este caso, 100 pasos).
    #tiempo_por_paso es el tiempo que debe esperar entre cada actualización del indicador.
    # Esto se obtiene dividiendo el tiempo total entre el número de pasos.
    tiempo_por_paso = tiempo_total / numero_de_pasos

#3. Iteramos sobre cada paso del indicador de carga.
    # Bucle para actualizar el indicador de carga desde 0% hasta 100%.
    for paso in range(numero_de_pasos + 1):
        # Calcula el porcentaje completado en el paso actual.
        porcentaje_actual = paso
        
#4. Imprimimos el porcentaje de carga en la misma línea de la consola.
        # El carácter \r mueve el cursor al principio de la línea para sobrescribir el texto anterior.
        sys.stdout.write(f"\rCARGA: {porcentaje_actual}%")
        
#5. Forzamos la actualización del texto en la consola.
        # Asegura que el texto se muestre de inmediato en la consola.
        sys.stdout.flush()

#6. Esperamos el tiempo calculado antes de actualizar al siguiente paso.
        # Pausa la ejecución del programa por el tiempo calculado antes de actualizar el próximo paso.
        # Esto crea el efecto de animación de carga.
        time.sleep(tiempo_por_paso)

#7. Una vez completada la carga, mostramos un mensaje final indicando que la carga está completa.
    # El carácter \r mueve el cursor al principio de la línea y el texto adicional borra cualquier residual de la línea anterior.
    sys.stdout.write("\rCARGA: 100% ")
    print("¡LA CARGA SE COMPLETÓ EXITOSAMENTE!\n")
    

#8. Forzamos la actualización del mensaje final en la consola.
    # Asegura que el mensaje final se muestre de inmediato en la consola.
    sys.stdout.flush()


#carga_completada(10)
    # Llama a la función con un tiempo total de 10 segundos
    # Esto hará que el indicador de carga se actualice cada 0.1 segundos (10 segundos / 100 pasos)
    # durante 10 segundos, mostrando el progreso desde 0% hasta 100%.


    
# Definimos una función de espera de 1 segundo para la selección del valor correspondiente ala ciudad
def pausa_ciudad_segundos():
    print("POR FAVOR, ESPERA MIENTRAS SE PROCESA TU SELECCIÓN DE CIUDAD...")
    time.sleep(1)
    

# Definimos una función de espera de 1 segundo para la selección del valor correspondiente al cine
def pausa_cine_segundos():
    print("POR FAVOR, ESPERA MIENTRAS SE PROCESA TU SELECCIÓN DE CINE...")
    time.sleep(1)



# Esta función intenta cerrar una ventana emergente que puede aparecer en el navegador.
# La función realiza los siguientes pasos:
    #1. Espera a que la ventana emergente sea visible en la página web.
    #2. Busca el botón de cerrar de la ventana emergente usando su identificador (ID).
    #3. Hace clic en el botón para cerrar la ventana emergente.
    #4. Imprime un mensaje indicando que se está cerrando la ventana emergente.
# Si ocurre algún error durante este proceso (como si el botón no se encuentra o no se puede hacer clic),
# se captura el error y se imprime un mensaje indicando el problema.
def cerrar_ventana_emergente(driver, espera):

    try:
        # Espera hasta que un elemento con la clase 'tk' esté presente en la página.
        espera.until(EC.presence_of_element_located((By.CSS_SELECTOR, "div.tk")))

        # Busca el botón para cerrar la ventana emergente usando su ID 'takeover-close'.
        boton_cerrar = driver.find_element(By.ID, "takeover-close")

        # Simula un clic en el botón para cerrar la ventana emergente.
        boton_cerrar.click()
        print("\nESTAMOS CERRANDO LA VENTANA EMERGENTE. POR FAVOR, ESPERE...")

        #Añade una pausa de cinco segungos mientras se cierra la ventana emergente
        time.sleep(5)
        print("¡LA VENTANA EMERGENTE SE HA CERRADO CON ÉXITO.!\n")

    except Exception as e:
        # Si ocurre un error al intentar cerrar la ventana, imprime un mensaje con los detalles del error.
        print(f"\nNO SE PUDO CERRAR LA VENTANA EMERGENTE: {e}\n")






#Esta función busca y cierra el segundo ítem (primer <li>) seleccionado en una lista de opciones visibles en la interfaz de usuario.

#En el contexto de la página web, la función realiza lo siguiente:
    #1. Identifica todos los ítems de lista (<li>) que están dentro de la lista de opciones seleccionadas (con la clase "chosen-choices").
    #2. Filtra estos ítems para obtener solo aquellos que tienen la clase "search-choice", que indica que son ítems seleccionados por el usuario.
    #3. Verifica que haya al menos dos ítems en la lista filtrada.
    #4. Selecciona el segundo ítem de la lista filtrada (índice 1, ya que los índices comienzan en 0).
    #5. Dentro de este ítem, busca el botón de cierre (etiqueta <a> con la clase "search-choice-close") y hace clic en él para cerrar o eliminar el ítem de la lista.

#Esta función es útil para manejar la interfaz de usuario donde se permite al usuario seleccionar múltiples opciones y luego eliminar algunas de ellas.

def eliminar_primer_item_li_lista():

    time.sleep(1)

# 1: Encontrar todos los elementos <li> dentro de la lista con la clase "chosen-choices"
    # En el CODIGO HTML, estos elementos <li> están dentro del <ul> con la clase "chosen-choices".
    # Esto selecciona todos los ítems de lista que ya han sido seleccionados (marcados) en la interfaz.
    elementos_lista = driver.find_elements(By.CSS_SELECTOR, 'ul.chosen-choices li')

# 2: Filtrar solo los elementos <li> que tienen la clase "search-choice"
    # De todos los elementos <li> encontrados, estamos filtrando solo aquellos que tienen la clase "search-choice".
    # En el CODIGO HTML, esto incluye los elementos que representan opciones seleccionadas por el usuario.
    elementos_filtrados = [elemento for elemento in elementos_lista if 'search-choice' in elemento.get_attribute('class')]

# 3: Asegurarse de que haya al menos dos elementos filtrados
    # Verificamos que haya al menos dos ítems en la lista filtrada para evitar errores si hay menos de dos.
    # Esto es importante porque queremos operar sobre el segundo ítem en la lista.
    if len(elementos_filtrados) > 1:

# 4: Seleccionar el segundo elemento (corresponde al primer <li> visible al usuario) en la lista filtrada.
        # La lista de ítems seleccionados comienza con índice 0, por lo que el segundo ítem tiene índice 1.
        # Aquí estamos seleccionando el segundo ítem de la lista filtrada (índice 1).
        segundo_elemento_primer_li = elementos_filtrados[1]
        
# 5: Cerrar el segundo ítem de la lista seleccionada
        # Dentro del ítem seleccionado, buscamos el botón de cierre (etiqueta <a> con la clase "search-choice-close").
        # Este botón se encuentra dentro del ítem de la lista y se usa para eliminar o cerrar ese ítem.
        # En el CODIGO HTML, este botón es el <a> dentro de <li> que tiene la clase "search-choice-close".
        boton_cerrar = segundo_elemento_primer_li.find_element(By.CSS_SELECTOR, 'a.search-choice-close')

        # Hacemos clic en este botón para eliminar el ítem de la lista de opciones seleccionadas.
        # El método `click()` simula un clic del usuario sobre el botón de cierre, lo que provoca que el ítem se elimine visualmente de la interfaz de usuario.
        boton_cerrar.click()
        # Mensaje en consola indicando que la operación fue exitosa
        print("¡EL PRIMER ÍTEM (<li>) DE LA LISTA HA SIDO CERRADO EXITOSAMENTE!\n")






#Esta función busca y cierra el tercer ítem (segundo <li>) seleccionado en una lista de opciones visibles en la interfaz de usuario.

#En el contexto de la página web, la función realiza lo siguiente:
    #1. Identifica todos los ítems de lista (<li>) que están dentro de la lista de opciones seleccionadas (con la clase "chosen-choices").
    #2. Filtra estos ítems para obtener solo aquellos que tienen la clase "search-choice", que indica que son ítems seleccionados por el usuario.
    #3. Verifica que haya al menos dos ítems en la lista filtrada.
    #4. Selecciona el tercer ítem de la lista filtrada (índice 2, ya que los índices comienzan en 0).
    #5. Dentro de este ítem, busca el botón de cierre (etiqueta <a> con la clase "search-choice-close") y hace clic en él para cerrar o eliminar el ítem de la lista.

#Esta función es útil para manejar la interfaz de usuario donde se permite al usuario seleccionar múltiples opciones y luego eliminar algunas de ellas.

def eliminar_segundo_item_li_lista():
    
    time.sleep(1)
    
# 1: Encontrar todos los elementos <li> dentro de la lista con la clase "chosen-choices"
    # En el CODIGO HTML, estos elementos <li> están dentro del <ul> con la clase "chosen-choices".
    # Esto selecciona todos los ítems de lista que ya han sido seleccionados (marcados) en la interfaz.
    elementos_lista = driver.find_elements(By.CSS_SELECTOR, 'ul.chosen-choices li')

# 2: Filtrar solo los elementos <li> que tienen la clase "search-choice"
    # De todos los elementos <li> encontrados, estamos filtrando solo aquellos que tienen la clase "search-choice".
    # En el CODIGO HTML, esto incluye los elementos que representan opciones seleccionadas por el usuario.
    elementos_filtrados = [elemento for elemento in elementos_lista if 'search-choice' in elemento.get_attribute('class')]

# 3: Asegurarse de que haya al menos dos elementos filtrados
    # Verificamos que haya al menos dos ítems en la lista filtrada para evitar errores si hay menos de dos.
    # Esto es importante porque queremos operar sobre el segundo ítem en la lista.
    if len(elementos_filtrados) > 1:

# 4: Seleccionar el tercer elemento (corresponde al segundo <li> visible al usuario) en la lista filtrada.
        # La lista de ítems seleccionados comienza con índice 0, por lo que el segundo ítem tiene índice 1
        # y el tercer ítem tiene índice 2. Aquí estamos seleccionando el tercer ítem de la lista filtrada.
        tercer_elemento_segundo_li = elementos_filtrados[2]
        
# 5: Cerrar el segundo ítem de la lista seleccionada
        # Dentro del ítem seleccionado, buscamos el botón de cierre (etiqueta <a> con la clase "search-choice-close").
        # Este botón se encuentra dentro del ítem de la lista y se usa para eliminar o cerrar ese ítem.
        # En el CODIGO HTML, este botón es el <a> dentro de <li> que tiene la clase "search-choice-close".
        boton_cerrar = tercer_elemento_segundo_li.find_element(By.CSS_SELECTOR, 'a.search-choice-close')

        # Hacemos clic en este botón para eliminar el ítem de la lista de opciones seleccionadas.
        # El método `click()` simula un clic del usuario sobre el botón de cierre, lo que provoca que el ítem se elimine visualmente de la interfaz de usuario.
        boton_cerrar.click()
        # Mensaje en consola indicando que la operación fue exitosa
        print("¡EL SEGUNDO ÍTEM (<li>) DE LA LISTA HA SIDO CERRADO EXITOSAMENTE!\n")




# Esta función elimina las filas incompletas de una hoja de cálculo de Excel.
# Revisa cada fila a partir de la segunda (asumiendo que la primera fila contiene encabezados) 
# y verifica si alguna celda en las primeras 8 columnas está vacía.
# Si encuentra celdas vacías en una fila, añade el número de esa fila a una lista de filas a eliminar.
# Luego, elimina las filas identificadas en orden inverso para evitar problemas con el cambio de índices durante la eliminación.
# Parámetros de la función.:
    #hoja_excel: Es el objeto que representa la hoja de cálculo en la que se trabajan los datos.

def eliminar_filas_incompletas(hoja_excel):

    time.sleep(1)

# 1: Crear una lista para guardar los números de las filas que vamos a eliminar.
    # Método `append`: Añade un elemento al final de una lista.
    filas_a_eliminar = []

# 2: Revisar cada fila de la hoja de cálculo, comenzando desde la segunda fila (Fila 2).
    # La primera fila en Excel suele tener encabezados (títulos de las columnas). Por eso, empezamos a revisar desde la segunda fila.
    # Revisa hasta la columna 'H' (columna número 8), para encontrar celdas vacías en esas columnas.
    # Método `iter_rows`: Itera sobre las filas de la hoja de cálculo dentro de un rango específico de filas y columnas.

# 2.1 Revisar cada fila desde la Fila 2 hasta la Columna 'H'
    for fila in hoja_excel.iter_rows(min_row=2, max_col=8):

# 3: Comprobar si alguna celda en la fila está vacía.
        # En cada fila, revisamos las celdas desde la Columna 'A' hasta la Columna 'H'. Si alguna celda está vacía (sin datos), consideramos que la fila está incompleta.
        # `any`: Función que devuelve True si al menos uno de los elementos en un iterable es True.

# 3.1 Verificar celdas vacías en las columnas de la A a la H
        if any(celda.value is None or celda.value == "" for celda in fila):

# 4: Añadir el número de la fila a la lista de filas a eliminar.
            # Método `append`: Añade el número de la fila a la lista `filas_a_eliminar`.
            filas_a_eliminar.append(fila[0].row)

# 4.5: Informar al usuario sobre las filas incompletas detectadas.
    if filas_a_eliminar:
        time.sleep(1)
        print("\n\nAL PARECER, HEMOS DETECTADO UNA SERIE DE DATOS COPIADOS INCOMPLETAMENTE.")
        time.sleep(1)
        print(f"SE HAN DETECTADO {len(filas_a_eliminar)} FILAS INCOMPLETAS.")
        time.sleep(1)
        print(f"LIMPIANDO LOS DATOS INCORRECTOS PARA GARANTIZAR LA EXACTITUD DE LA INFORMACIÓN...\n")
        time.sleep(1)
# 5: Itera sobre los números de fila en la lista `filas_a_eliminar` en orden inverso.
    # Esto es necesario para evitar problemas con el cambio de índices de las filas mientras se eliminan.
    # Es más seguro eliminar las filas empezando desde la última hacia la primera.
    # De esta forma, no cambiamos los números de las filas que aún no hemos borrado.
    # Método `reversed`: Devuelve un iterador que accede a los elementos en orden inverso.

# 5.1 Eliminar filas en orden inverso
        for fila in reversed(filas_a_eliminar):

# 6: Eliminar la fila específica de la hoja de cálculo.
    # Elimina una fila en la hoja de cálculo usando el método `delete_rows`.
    # `hoja_excel.delete_rows(fila, 1)` elimina filas de la hoja de cálculo.
        # El primer parámetro (`fila`) es el número de la fila que se va a eliminar.
        # El segundo parámetro (`1`) indica cuántas filas se eliminarán a partir de la fila especificada.
        # En este caso, `1` significa que se eliminará solo una fila.
            hoja_excel.delete_rows(fila, 1)

        # Informar al usuario que las filas han sido eliminadas.
        print(f"¡SE HAN CORREGIDO LOS DATOS INCOMPLETOS DE MANERA ADECUADA!\n\n")
    else:
        # Si no hay filas incompletas
        print("NO SE DETECTARON FILAS INCOMPLETAS.\n\n")









#Esta función se encarga de seleccionar una ciudad y un cine en una página web utilizando los menús desplegables disponibles.
    #Pasos que realiza:
        #Espera a que aparezca el menú de selección de ciudades:
            #Se asegura de que el menú desplegable para elegir una ciudad esté visible en la página web.
        #Selecciona la ciudad deseada:
            #Elige la ciudad que se especifica mediante el valor proporcionado.
        #Espera a que aparezca el menú de selección de cines:
            #Se asegura de que el menú desplegable para elegir un cine esté disponible después de seleccionar una ciudad.
        #Selecciona el cine deseado:
            #Elige el cine que se especifica mediante el valor proporcionado.
        #Formatea el nombre del cine:
            #Convierte el nombre del cine a mayúsculas y reemplaza guiones bajos por espacios para una presentación más clara.
        #Devuelve el nombre del cine formateado:
            #Proporciona el nombre del cine ya ajustado para su uso en otras partes del programa.
    #Parámetros:
        #driver: El objeto del navegador web que controla la automatización de la página.
        #espera: Un objeto que permite esperar a que los elementos de la página estén disponibles.
        #valor_ciudad: El valor del menú desplegable que corresponde a la ciudad que se quiere seleccionar.
        #valor_cine: El valor del menú desplegable que corresponde al cine que se quiere seleccionar.
    #Retorna:
        #nombre_cine: El nombre del cine seleccionado, formateado en mayúsculas y con espacios en lugar de guiones bajos.

def elegir_ciudad_cine(driver, espera, valor_ciudad, valor_cine):

    # Pausa la ejecución del script para permitir que la página web se cargue completamente antes de interactuar.
    pausa_ciudad_segundos()
    carga_completada(2)
    
# Proceso 1: Seleccionar Ciudad
    # Esperamos a que el menú para seleccionar la ciudad aparezca en la página web.
        # 'espera' es un objeto WebDriverWait que permite esperar de manera explícita hasta que el elemento esté presente.
        # Se utiliza 'EC.presence_of_element_located' para esperar hasta que el elemento con el ID 'cmbCiudades' esté en el DOM.
        # 'By.ID' indica que se está buscando un elemento por su atributo 'id'.
        # Este menú es un elemento HTML de tipo <select> con el ID "cmbCiudades".
        # En HTML, se vería algo así: <select id="cmbCiudades">...</select>
    VARIABLE_HTML_select_ciudad = espera.until(EC.presence_of_element_located((By.ID, "cmbCiudades")))

    # Una vez que el elemento está disponible, se crea un objeto Select que facilita la interacción con menús desplegables (select elements).
        # Usamos una herramienta llamada 'Select' para trabajar con este menú desplegable.
        # Esta herramienta nos permite seleccionar opciones dentro del menú desplegable.
        # El menú tiene varias opciones representadas por elementos <option> dentro de él.
    select_ciudad = Select(VARIABLE_HTML_select_ciudad)
    
    # Seleccionamos la ciudad que queremos usar.
        # Selecciona la ciudad deseada en el menú desplegable usando el valor proporcionado ('valor_ciudad').
        # 'select_by_value' selecciona una opción basada en el atributo 'value' del elemento <option>.
        # En HTML, esto corresponde a: <option value="valor_ciudad">Ciudad</option>
        # 'valor_ciudad' es el identificador de la ciudad que queremos elegir, que corresponde al valor del atributo 'value' en el <option>.
        # Por ejemplo, si queremos elegir "Ciudad A", el HTML de la opción sería: <option value="0">Ciudad A</option>
    select_ciudad.select_by_value(valor_ciudad)


# Proceso 2: Seleccionar Cine
    # Espera hasta que las opciones del menú desplegable de cines estén disponibles.
        # Esto asegura que todas las opciones de cine se hayan cargado completamente antes de intentar seleccionar una.
        # Después de seleccionar una ciudad, debemos esperar a que el menú de selección de cines aparezca.
        # Esto es necesario porque el menú de cines solo se carga después de elegir una ciudad.
        # El selector CSS 'select#cmbComplejos option' se usa para esperar hasta que todas las opciones dentro del select con ID 'cmbComplejos' estén presentes.
        # En HTML, este menú es otro elemento <select> con el ID "cmbComplejos".
    espera.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, "select#cmbComplejos option")))

    # Pausa la ejecución del script para permitir que la página web se cargue completamente antes de interactuar.
    pausa_cine_segundos()
    carga_completada(2)
    
    # Busca el elemento del menú desplegable de cines en la página web usando su ID.
        # 'espera.until' asegura que el elemento esté disponible antes de continuar.
        # Esperamos a que el menú de cines esté disponible y visible en la página.
        # El ID 'cmbComplejos' corresponde al menú desplegable de cines.
        # Queremos asegurarnos de que el menú donde elegimos el cine esté completamente cargado antes de interactuar con él.
    VARIABLE_HTML_select_cine = espera.until(EC.presence_of_element_located((By.ID, "cmbComplejos")))
    
    # Usamos la herramienta 'Select' de nuevo para trabajar con el menú desplegable de cines.
        # Crea un objeto Select para manejar la selección dentro del menú desplegable de cines.
        # Similar al menú de ciudades, se utiliza para seleccionar una opción basada en su valor.
        # Esto nos permite seleccionar una opción en el menú de cines, que también es un <select> en HTML.
    select_cine = Select(VARIABLE_HTML_select_cine)
    
    # Elegimos el cine que queremos de la lista desplegable.
        # Selecciona el cine deseado en el menú desplegable usando el valor proporcionado ('valor_cine').
        # 'select_by_value' selecciona una opción en el menú de cines basada en el atributo 'value' del elemento <option>.
        # En HTML, esto corresponde a: <option value="valor_cine">Cine</option>
        # 'valor_cine' es el identificador del cine que queremos seleccionar, que corresponde al valor del atributo 'value' en el <option>.
        # Por ejemplo, si queremos seleccionar "Cine B", el HTML de la opción sería: <option value="cinepolis-federal-mall">Cine B</option>
    select_cine.select_by_value(valor_cine)
    

# Proceso 3: Formatear y Devolver Nombre del Cine
    # Convertimos el nombre del cine a mayúsculas y reemplazamos los guiones con espacios.
        # 'valor_cine.upper()' convierte todo el texto a mayúsculas para uniformidad.
        # 'replace("-", " ")' reemplaza guiones bajos con espacios para hacer el nombre más legible.
        # Esto hace que el nombre del cine se vea más limpio y sea más fácil de leer.
        # Por ejemplo, si 'valor_cine' es "cinepolis-federal-mall", lo convertimos a "CINEPOLIS FEDERAL MALL".
    nombre_cine = valor_cine.upper().replace("-", " ")
    
    # Devolvemos el nombre del cine ya formateado.
        # Devuelve el nombre del cine en el formato legible.
        # Este valor es el que puede ser utilizado en otras partes del programa o para mostrar al usuario.
    return nombre_cine





# Esta función copia datos sobre películas y horarios desde una fuente web a una hoja de cálculo de Excel.
# Toma como entrada el nombre del cine (`nombre_cine`) y un objeto que representa el archivo de Excel (`libro_excel`).
# La función realiza las siguientes tareas:
    #1.Preparar Nombre de la Hoja:
        #Crea una nueva hoja en el archivo de Excel con un nombre basado en el nombre del cine.
        #Limita el nombre de la hoja a un máximo de 31 caracteres y asegura que el nombre sea único dentro del archivo.
    #2.Extraer y Organizar Datos:
        #Recolecta los nombres de las películas y sus formatos desde la página web.
        #Organiza los datos en listas y diccionarios para facilitar la escritura en la hoja de Excel.
    #3.Escribir Nombres de Películas:
        #Escribe los nombres de las películas en la columna "TITULO" (columna E) de la hoja de Excel.
    #4.Extraer Horarios y Formatos:
        #Busca y almacena los horarios de proyección y los formatos asociados para cada película desde la fuente web.
        #Organiza esta información en diccionarios, diferenciando el formato y el horario de cada película.
    #5.Escribir Horarios y Formatos:
        #Registra la información de horarios y formatos en las columnas correspondientes de la hoja de Excel:
            #Columna D: Nombre del cine.
            #Columna E: Títulos de las películas.
            #Columna F: Horas de proyección.
            #Columna G: Idioma o formato principal (como 3D o DIG).
            #Columna H: Formatos adicionales de la película (como subtitulado o doblado).
        #Asegura que si no se especifica un formato especial, se use "2D" como formato predeterminado.
# La función organiza y almacena la información de manera estructurada en Excel para facilitar su análisis y consulta posterior.

def copiar_datos_a_excel(nombre_cine, libro_excel):

# 1. Normalización del nombre del cine para el nombre de la hoja
    # Preparar el nombre de la hoja en el libro de Excel
    # El nombre de la hoja debe ser único y no debe exceder los 31 caracteres
    # Convertimos el nombre a letras Mayusculas para una mejor vista al usuario en las hojas de excel
    nombre_cine_corto = (nombre_cine[:31] if len(nombre_cine) > 31 else nombre_cine).upper()




# 1.1. Ajuste del nombre de la hoja si ya existe una hoja con ese nombre en el libro
    # Si el nombre del cine es más largo de 31 caracteres, lo acortamos a los primeros 31 caracteres.
    # También convertimos el nombre a mayúsculas para uniformidad.

    while nombre_cine_corto in libro_excel.sheetnames:
        # Si ya existe una hoja con el mismo nombre en el libro de Excel, modificamos el nombre
        # para que sea único añadiendo un número al final.
        nombre_cine_corto = (nombre_cine_corto[:30] + '_' + str(len(libro_excel.sheetnames))).upper()
        # Acortamos el nombre a 30 caracteres y añadimos un número para asegurar que el nombre sea único.
        # Luego, convertimos el nombre a mayúsculas.




# 1.2. Creación de una nueva hoja con el nombre ajustado
    #Crear una nueva hoja en el libro de Excel
    hoja_excel = libro_excel.create_sheet(title=nombre_cine_corto)
    # Creamos una nueva hoja con el nombre preparado y la añadimos al libro de Excel.



# 2. Inicialización de listas y diccionarios para almacenar datos
    lista_peliculas = []
    # Creamos una lista vacía para almacenar los nombres de las películas que encontraremos.

    lista_formatos_por_hora = {}
    # Creamos un diccionario vacío para almacenar los formatos de las películas según los horarios.

    formato_por_pelicula = {}
    # Creamos un diccionario vacío para almacenar los formatos disponibles para cada película.


    # Extraer películas y formatos
    # Este bloque de código busca en la página web información sobre las películas que se muestran
    # y los formatos en los que están disponibles, y guarda esta información en listas y diccionarios.



# 3. Extracción de títulos de películas y formatos de la página web
    #Buscar todos los elementos que contienen información sobre las películas
    for VARIABLE_HTML_div in driver.find_elements(By.CSS_SELECTOR, "div.descripcion.col8"):
        # Aquí, `driver.find_elements` busca todos los bloques de información que contienen detalles
        # de las películas. Estos bloques están marcados con una clase específica llamada "descripcion col8".


# 3.1. Obtención del título de la película
    #Encontrar el título de la película dentro de cada bloque
        VARIABLE_HTML_h3 = VARIABLE_HTML_div.find_element(By.TAG_NAME, "h3")
        # Dentro de cada bloque de película, buscamos un título que está dentro de una etiqueta HTML <h3>.

        VARIABLE_HTML_a = VARIABLE_HTML_h3.find_element(By.TAG_NAME, "a")
        # El título de la película se encuentra dentro de un enlace HTML <a>.

        pelicula = VARIABLE_HTML_a.text
        # Extraemos el texto del enlace, que es el nombre de la película.

        lista_peliculas.append(pelicula)
        # Agregamos el nombre de la película a una lista llamada `lista_peliculas`.


# 3.2. Extracción de formatos disponibles para la película
    #Buscar los formatos disponibles para la película en imágenes
        img_tags = VARIABLE_HTML_div.find_elements(By.TAG_NAME, "img")
        # Dentro del bloque de la película, buscamos todas las imágenes (<img>) que indican los formatos disponibles.

    #Extraer la información del formato de las imágenes
        formatos = [img.get_attribute("alt") for img in img_tags if img.get_attribute("alt")]
        # Para cada imagen, extraemos el texto que describe el formato (como "3D" o "DIG") que se encuentra en el atributo "alt".
        # Solo guardamos formatos si el atributo "alt" no está vacío.

        formato_por_pelicula[pelicula] = formatos
        # Guardamos la lista de formatos para cada película en un diccionario llamado `formato_por_pelicula`.
        # La clave del diccionario es el nombre de la película, y el valor es la lista de formatos disponibles.



# 4. Escritura de títulos de películas en la hoja de Excel
    # Encabezado de columna para títulos de películas
    hoja_excel['E1'] = "TITULO"
    # En la primera fila de la columna 'E' de la hoja de Excel, escribimos el título "TITULO".
    # Esto sirve como encabezado para la columna que contendrá los nombres de las películas.
    # En Excel, las columnas están etiquetadas con letras, y 'E' representa la columna en la que se guardarán los títulos.

    for idx, pelicula in enumerate(lista_peliculas, start=2):
        # Iteramos sobre la lista de nombres de películas (`lista_peliculas`).
        # `enumerate` proporciona un índice (que comienza en 2) y el nombre de la película.
        # `start=2` significa que la numeración de filas comenzará en la fila 2 (porque la fila 1 se usa para encabezados).

        hoja_excel[f'E{idx}'] = pelicula
        # Para cada película, escribimos el nombre en la columna 'E' y en la fila correspondiente (`idx`).
        # Por ejemplo, la primera película se escribirá en la celda E2, la segunda en E3, etc.



# 5. Extracción de horas y formatos de las películas
    #Creamos un diccionario vacío llamado `lista_horas_por_pelicula`.
    lista_horas_por_pelicula = {}
    # Este diccionario se utilizará para almacenar las horas en las que se muestran las películas,
    # organizadas por película. La clave del diccionario será el nombre de la película,
    # y el valor será una lista de horas en las que se muestra esa película.


#Intentamos ejecutar el siguiente bloque de código, que extrae información sobre horarios de películas.
    try:
        for VARIABLE_HTML_div in driver.find_elements(By.CSS_SELECTOR, "div.descripcion.col8"):



# 5.1. Obtención del título de la película
            # Buscamos todos los bloques de información sobre películas en la página web.
            # Estos bloques están etiquetados con una clase específica llamada "descripcion col8".

            VARIABLE_HTML_h3 = VARIABLE_HTML_div.find_element(By.TAG_NAME, "h3")
            # Dentro de cada bloque, buscamos una etiqueta HTML <h3> que contiene el título de la película.

            VARIABLE_HTML_a = VARIABLE_HTML_h3.find_element(By.TAG_NAME, "a")
            # Dentro de la etiqueta <h3>, buscamos un enlace HTML <a> que contiene el nombre de la película.

            pelicula = VARIABLE_HTML_a.text
            # Extraemos el texto del enlace, que es el nombre de la película.



# 5.2. Obtención de los horarios de la película
            divs_horarios = VARIABLE_HTML_div.find_elements(By.CSS_SELECTOR, "div.horarioExp")
            # Dentro del bloque de información de la película, buscamos todos los elementos que contienen horarios.
            # Estos elementos están etiquetados con una clase llamada "horarioExp".

            lista_horas = []
            # Creamos una lista vacía llamada `lista_horas` para guardar los horarios en los que se muestra la película.

            for div_horario in divs_horarios:
                # Recorremos cada bloque de horarios (`div_horario`) que hemos encontrado para la película.
                # Cada uno de estos bloques contiene información sobre los horarios y formatos de proyección.



# 5.2.1. Obtención del formato de la película
                VARIABLE_HTML_formato = div_horario.find_element(By.CSS_SELECTOR, "div.col3.cf.ng-binding")
                # Dentro de cada bloque de horarios, buscamos el elemento que contiene el formato de proyección.
                # Este elemento está etiquetado con una clase específica llamada "col3 cf ng-binding".

                formato = VARIABLE_HTML_formato.text.strip()
                # Extraemos el texto de este elemento, que indica el formato de proyección (como "2D", "3D").
                # `strip()` se usa para eliminar cualquier espacio en blanco innecesario al principio o al final del texto.

                etiquetas_html_p = VARIABLE_HTML_formato.find_elements(By.TAG_NAME, "p")
                # Dentro del elemento de formato, buscamos todas las etiquetas HTML `<p>`, que pueden contener información adicional.
                # En algunos casos, el formato puede estar dividido en múltiples partes dentro de estas etiquetas `<p>`.

                for etiqueta_html_p in etiquetas_html_p:
                    # Recorremos cada etiqueta HTML <p> dentro del elemento de formato.
                    # Cada etiqueta <p> puede contener partes del formato de proyección.

                    etiquetas_html_span = etiqueta_html_p.find_elements(By.TAG_NAME, "span")
                    # Dentro de cada etiqueta <p>, buscamos todas las etiquetas HTML <span>.
                    # Las etiquetas <span> pueden contener partes individuales del formato de proyección.

                    if len(etiquetas_html_span) == 2:
                        # Si encontramos exactamente dos etiquetas <span> dentro de una etiqueta <p>:
                        # Esto suele significar que el formato de proyección está dividido en dos partes (como "2D" y "Digital").
                        
                        formato = f"{etiquetas_html_span[0].text} {etiquetas_html_span[1].text}"
                        # Unimos el texto de las dos etiquetas <span> en una sola cadena.
                        # Esto da como resultado un formato completo (por ejemplo, "2D Digital").



# 5.2.2. Obtención de las horas de la película
                VARIABLE_HTML_tiempos = div_horario.find_elements(By.CSS_SELECTOR, "time.btn.btnhorario.ng-scope")
                # Dentro del bloque de horarios, buscamos todos los elementos que representan los tiempos de proyección.
                # Estos elementos están etiquetados con una clase específica llamada "btn btnhorario ng-scope" en una etiqueta <time>.

                for VARIABLE_HTML_tiempo in VARIABLE_HTML_tiempos:
                    # Recorremos cada elemento que representa un horario de proyección.
                    # Estos elementos han sido previamente identificados y almacenados en `VARIABLE_HTML_tiempos`.

                    VARIABLE_HTML_a = VARIABLE_HTML_tiempo.find_element(By.TAG_NAME, "a")
                    # Dentro de cada elemento de horario, buscamos una etiqueta HTML <a> que contiene el tiempo específico de proyección.

                    hora = VARIABLE_HTML_a.text
                    # Extraemos el texto de esta etiqueta <a>, que indica el horario (por ejemplo, "14:00").

                    lista_horas.append(hora)
                    # Añadimos este horario a la lista llamada `lista_horas`, que mantiene todos los horarios disponibles para esta película.

                    lista_formatos_por_hora[hora] = formato
                    # Asignamos el formato de proyección (como "2D" o "3D Digital") a este horario específico en el diccionario `lista_formatos_por_hora`.
                    # Esto crea una relación entre cada horario y su formato correspondiente.

            lista_horas_por_pelicula[pelicula] = lista_horas
            # Finalmente, añadimos la lista de horarios (`lista_horas`) para esta película a un diccionario llamado `lista_horas_por_pelicula`.
            # Esto guarda todos los horarios disponibles para la película en el diccionario bajo la clave que es el nombre de la película.

    except Exception as e:
        # Este bloque se ejecuta si ocurre algún error durante la ejecución del código que intenta extraer los horarios y formatos de las películas.
        # La palabra clave `except` captura cualquier tipo de error que pueda ocurrir en el bloque de código anterior (donde se extraen los horarios y formatos).

        print(f"Error al extraer horas y formatos de las películas: {e}")
        # Si ocurre un error, se imprime un mensaje que indica que hubo un problema al intentar extraer la información.
        # El mensaje incluye detalles sobre el error específico, que se almacenan en la variable `e`.
        # `e` contiene la descripción del error, lo que ayuda a identificar qué salió mal y facilita la resolución del problema.




# 6. Escritura de datos de horarios y formatos en la hoja de Excel
    #Estas líneas establecen los encabezados de las columnas en la primera fila de la hoja de Excel.    
    # Los encabezados son etiquetas que describen qué tipo de información contiene cada columna:
    hoja_excel['F1'] = "HORA"           # - Columna F: "HORA" (para los horarios de proyección)
    hoja_excel['G1'] = "IDIOMA"         # - Columna G: "IDIOMA" (para el idioma del formato)
    hoja_excel['H1'] = "FORMATO"        # - Columna H: "FORMATO" (para el formato de la película, como "2D" o "3D")
    hoja_excel['D1'] = "NOMBRE CINE"    # - Columna D: "NOMBRE CINE" (para el nombre del cine)

    # Comienza en la segunda fila para los datos
    fila = 2
    # Inicia la variable `fila` en 2 porque la primera fila (1) ya está reservada para los encabezados de las columnas.
    # La variable `fila` se usa para controlar en qué fila se debe escribir la información de cada película.

    for pelicula, lista_horas in lista_horas_por_pelicula.items():
        # Se recorre el diccionario `lista_horas_por_pelicula`, donde cada clave es el nombre de una película y el valor es una lista de horarios para esa película.
        # `pelicula` es el nombre de la película, y `lista_horas` es la lista de horarios para esa película.

        for hora in lista_horas:
            # Se recorre cada horario en la lista de horarios de la película actual.
            # `hora` representa un horario específico para la película en formato de texto (por ejemplo, "14:00").

            hoja_excel[f'D{fila}'] = nombre_cine
            # Escribe el nombre del cine en la columna D de la fila actual.
            # `nombre_cine` es el nombre del cine donde se están viendo las proyecciones.

            hoja_excel[f'E{fila}'] = pelicula
            # Escribe el nombre de la película en la columna E de la fila actual.

            hoja_excel[f'F{fila}'] = hora
            # Escribe el horario de proyección en la columna F de la fila actual.

            formato_idioma = lista_formatos_por_hora.get(hora, "")
            # Obtiene el formato asociado al horario actual. `lista_formatos_por_hora` es un diccionario donde la clave es el horario y el valor es el formato.
            # Si el horario no se encuentra en el diccionario, se usa una cadena vacía como valor predeterminado.

            hoja_excel[f'G{fila}'] = formato_idioma
            # Escribe el formato del horario en la columna G de la fila actual.




# 6.1. Ajusta el formato en función del tipo (3D, DIG, 2D)
            if "3D" in formato_idioma or "DIG" in formato_idioma:
                # Este bloque de código verifica si el formato de la película contiene "3D" o "DIG".
                # "3D" y "DIG" son abreviaturas comunes para formatos especiales de películas, como 3D o Digital.

                hoja_excel[f'H{fila}'] = "3D" if "3D" in formato_idioma else "DIG"
                # Si "3D" está en el formato, se escribe "3D" en la columna H de la fila actual.
                # Si "DIG" está en el formato, se escribe "DIG" en la columna H de la fila actual.
                # Solo se coloca uno de los dos valores en la columna H, dependiendo de cuál esté presente en el formato.

                formato_idioma = formato_idioma.replace("3D", "").replace("DIG", "").strip()
                # Elimina "3D" o "DIG" del formato original.
                # `replace("3D", "")` elimina todas las apariciones de "3D" en el texto.
                # `replace("DIG", "")` hace lo mismo con "DIG".
                # `strip()` elimina los espacios adicionales que pueden quedar después de eliminar "3D" o "DIG".

                hoja_excel[f'G{fila}'] = formato_idioma
                # Después de limpiar el formato, escribe el texto restante en la columna G de la fila actual.
                # Esto asegura que solo el formato principal (como "2D") quede en la columna G después de quitar "3D" o "DIG".

            
            if not hoja_excel[f'H{fila}'].value:
                # Verifica si la celda en la columna H (que contiene el formato especial) está vacía.
                # `hoja_excel[f'H{fila}'].value` obtiene el valor de la celda en la columna H para la fila actual.
                # Si no hay ningún valor (es decir, si la celda está vacía), significa que no se ha encontrado un formato especial como "3D" o "DIG".

                hoja_excel[f'H{fila}'] = "2D"
                # Si la celda en la columna H está vacía, se coloca "2D" en la celda.
                # Esto asegura que, si no se especifica ningún formato especial, el formato predeterminado "2D" se asigna automáticamente.
                # "2D" es un formato estándar para películas que no tienen un formato especial como 3D o Digital.




# 6.2. Agrega formatos adicionales de la película
            formatos_adicionales = formato_por_pelicula.get(pelicula, [])
            # Obtiene una lista de formatos adicionales para la película actual.
            # `formato_por_pelicula` es un diccionario donde cada clave es el nombre de una película y el valor es una lista de formatos adicionales.
            # `formato_por_pelicula.get(pelicula, [])` busca los formatos adicionales para la película en cuestión.
            # Si la película no tiene formatos adicionales en el diccionario, devuelve una lista vacía `[]`.

            
            if formatos_adicionales:
                # Verifica si hay formatos adicionales para la película.
                # `formatos_adicionales` es una lista de formatos que se recuperó para la película actual.
                # Si esta lista no está vacía, significa que hay formatos adicionales que deben agregarse a la celda de formato.

                formato_existente = hoja_excel[f'H{fila}'].value or ""
                # Obtiene el valor actual de la celda en la columna H para la fila actual.
                # `hoja_excel[f'H{fila}'].value` devuelve el valor en la celda.
                # Si la celda ya tiene un valor, `formato_existente` lo guardará; si está vacía, se asigna una cadena vacía `""`.

                hoja_excel[f'H{fila}'] = " ".join([f"{formato}" for formato in formatos_adicionales] + [formato_existente])
                # Une todos los formatos adicionales en una sola cadena de texto, separados por espacios.
                # `[f"{formato}" for formato in formatos_adicionales]` crea una lista de los formatos adicionales como cadenas de texto.
                # Se agrega `formato_existente` (el valor que ya estaba en la celda) al final de esta lista.
                # `join()` combina estos valores en una única cadena, separada por espacios.
                # Esta cadena resultante se asigna a la celda en la columna H, asegurando que todos los formatos adicionales y el formato existente se muestren.

            fila += 1
            # Incrementa el número de fila en 1.
            # Esto mueve la posición hacia abajo en la hoja de Excel para la siguiente película o conjunto de datos.
            # Así, cada película y sus formatos se colocan en filas sucesivas de la hoja de Excel.





# 7. Añade la fecha actual a la columna A de todas las filas
    #Establece el título de la primera columna en la hoja de Excel como 'FECHA'.
    hoja_excel['A1'] = 'FECHA'
    
    # Obtiene el número total de filas en la hoja de Excel que contienen datos.
    max_filas = hoja_excel.max_row
    
    # Obtiene la fecha actual en formato 'YYYY-MM-DD' (por ejemplo, '2024-08-12').
    fecha_actual = datetime.now().strftime('%Y-%m-%d')

    # Recorre todas las filas de la hoja de Excel, comenzando desde la fila 2 hasta la última fila con datos.
    for fila in range(2, max_filas + 1):
        # Establece el valor de la celda en la columna A (FECHA) para la fila actual con la fecha actual.
        hoja_excel[f'A{fila}'] = fecha_actual






# 8. Determina el país basado en la URL
    #Creamos un diccionario llamado 'dominio_a_pais'
    # Un diccionario es una estructura de datos que guarda pares de clave-valor.
    # En este caso, la clave es una extensión de dominio (como 'sv') y el valor es el nombre del país correspondiente (como 'EL SALVADOR').
    dominio_a_pais = {
        'sv': 'EL SALVADOR',      # La clave 'sv' se asocia con el valor 'EL SALVADOR'.
        'gt': 'GUATEMALA',        # La clave 'gt' se asocia con el valor 'GUATEMALA'.
        'cr': 'COSTA RICA',       # La clave 'cr' se asocia con el valor 'COSTA RICA'.
        'pa': 'PANAMÁ'            # La clave 'pa' se asocia con el valor 'PANAMÁ'.
    }


# Esta función determina el país al que pertenece una URL basada en el dominio de la página web.
# Toma como entrada una cadena de texto (`url`) que representa la dirección web.
    # La función realiza los siguientes pasos:
    # 1. Extrae la parte del dominio de la URL después de "https://" y antes del primer "/".
    # 2. Identifica la extensión del dominio, que es la parte final del dominio después del último punto (por ejemplo, "sv" en "cinepolis.com.sv").
    # 3. Usa un diccionario (`dominio_a_pais`) para mapear esta extensión de dominio a un nombre de país específico.
    # 4. Devuelve el nombre del país asociado a la extensión del dominio. Si la extensión no está en el diccionario, devuelve "DESCONOCIDO".
# Esta función es útil para clasificar o identificar la ubicación geográfica basada en el dominio de la página web.

    def obtener_pais(url):

        # Extrae el dominio de la URL, eliminando el protocolo (http://, https://) y cualquier ruta adicional.
        dominio = url.split('//')[-1].split('/')[0]
        # Obtiene la extensión del dominio, por ejemplo, "sv" de "cinepolis.com.sv".
        extension = dominio.split('.')[-1]
        # Busca la extensión en el diccionario para obtener el país. Si no se encuentra, devuelve 'DESCONOCIDO'.
        return dominio_a_pais.get(extension, 'DESCONOCIDO')

    # Obtiene el nombre del país basándose en la URL.
    nombrePais = obtener_pais(url)




# 9. Añade el país a la columna B de todas las filas
    # Establece el encabezado "PAIS" en la celda B1 de la hoja de Excel.
    hoja_excel["B1"] = "PAIS"
    # Recorre todas las filas con datos en la hoja de Excel, comenzando desde la fila 2.
    for fila in range(2, max_filas + 1):
        # Establece el valor de la celda en la columna B (PAIS) para la fila actual con el nombre del país.
        hoja_excel[f"B{fila}"] = nombrePais




# 10. Añade el nombre del cine a la columna C de todas las filas
    #Extrae el nombre del cine de la URL y lo convierte a mayúsculas.
    # Supone que la URL tiene el formato "https://nombre-del-cine.com.pa/..." 
    cadena_de_cine = url.split("https://")[1].split(".com.pa/")[0].upper()

    # Establece el encabezado "CINE" en la celda C1 de la hoja de Excel.
    hoja_excel["C1"] = "CINE"

    # Recorre todas las filas con datos en la hoja de Excel, comenzando desde la fila 2.
    for fila in range(2, max_filas + 1):
        # Establece el valor de la celda en la columna C (CINE) para la fila actual con el nombre del cine extraído.
        hoja_excel[f"C{fila}"] = cadena_de_cine

    #Eliminaremos las filas que fueron copiadas incompletamente
    eliminar_filas_incompletas(hoja_excel)




# Esta función elimina iconos o símbolos no deseados de la columna H en el archivo Excel.
# La columna H se selecciona dentro del archivo de Excel que se pasa como argumento (`libro_excel`).
# La función realiza lo siguiente:
    #1. Accede a la hoja activa del libro de Excel.
    #2. Recorre cada celda en la columna H.
    #3. Reemplaza cualquier icono o símbolo específico con una cadena vacía, dejando solo el texto limpio en la celda.
# Este proceso es útil para limpiar los datos antes de analizarlos o guardarlos, asegurando que solo se mantenga el texto relevante.

def eliminar_icono_de_columna_h(libro_excel):

    # Recorremos cada hoja en el archivo de Excel.
    # Cada hoja en un libro de Excel es una página donde se pueden encontrar varias tablas o datos.
    for hoja in libro_excel.worksheets:
        # Iteramos sobre cada fila en la columna H (la columna número 8), comenzando desde la segunda fila.
        # Usamos `min_col=8` y `max_col=8` para centrarnos solo en la columna H.
        # `min_row=2` significa que comenzamos en la segunda fila, generalmente porque la primera fila contiene encabezados.
        for fila in hoja.iter_rows(min_col=8, max_col=8, min_row=2):
            # `fila` es una lista que contiene las celdas de la columna H para una fila específica.
            # Tomamos la primera celda en esta lista, que corresponde a la columna H.
            celda = fila[0]
            # Verificamos si la celda tiene un valor y si ese valor es una cadena de texto.
            # `celda.value` contiene el contenido de la celda, y `isinstance(celda.value, str)` asegura que es texto.
            if celda.value and isinstance(celda.value, str):
                # `replace("icono-", "")` busca la palabra "icono-" en el texto y la reemplaza con una cadena vacía.
                # Esto elimina "icono-" del contenido de la celda si está presente.
                celda.value = celda.value.replace("icono-", "")



# Esta función elimina palabras repetidas en las celdas de la columna H en un archivo Excel.
# La columna H se selecciona dentro del archivo de Excel que se pasa como argumento (`libro_excel`).
# La función realiza los siguientes pasos:
    # 1. Accede a la hoja activa del libro de Excel.
    # 2. Recorre cada celda en la columna H.
    # 3. En cada celda, identifica y elimina las palabras que se repiten.
    # 4. Guarda el contenido limpio, con palabras únicas, en la misma celda.
# Este proceso ayuda a limpiar los datos en la columna H para que no haya duplicados y sea más fácil analizarlos.

def eliminar_palabras_repetidas_columna_h(libro_excel):

    # Recorremos cada hoja en el archivo de Excel.
    # Esto es necesario porque queremos aplicar la limpieza a todas las hojas en el libro.
    for hoja in libro_excel.worksheets:
        # Iteramos sobre cada fila en la columna H (la columna número 8), comenzando desde la segunda fila.
        # Usamos `min_col=8` y `max_col=8` para centrarnos solo en la columna H.
        # `min_row=2` se usa porque la primera fila generalmente contiene encabezados.
        for fila in hoja.iter_rows(min_col=8, max_col=8, min_row=2):
            # `fila` es una lista que contiene las celdas de la columna H para una fila específica.
            # Tomamos la primera celda en esta lista, que corresponde a la columna H.
            celda = fila[0]
            # Verificamos si la celda tiene un valor y si ese valor es una cadena de texto.
            # `celda.value` contiene el contenido de la celda, y `isinstance(celda.value, str)` asegura que sea texto.
            if celda.value and isinstance(celda.value, str):
                # Dividimos el texto en palabras usando el espacio como separador.
                palabras = celda.value.split()
                # Usamos `dict.fromkeys(palabras)` para eliminar palabras duplicadas.
                # Convertimos la lista de palabras en un diccionario (donde las claves son únicas) y luego de vuelta a una lista.
                palabras_unicas = list(dict.fromkeys(palabras))
                # Volvemos a unir las palabras únicas en una cadena de texto, separadas por espacios.
                celda.value = ' '.join(palabras_unicas)



# Esta función convierte todo el texto en la columna H de un archivo Excel a letras mayúsculas.
# La columna H se selecciona dentro del archivo de Excel que se pasa como argumento (`libro_excel`).
# La función realiza los siguientes pasos:
    # 1. Accede a la hoja activa del libro de Excel.
    # 2. Recorre cada celda en la columna H.
    # 3. Toma el texto de cada celda y lo convierte a letras mayúsculas.
    # 4. Guarda el texto convertido en la misma celda.
# Esto asegura que todo el texto en la columna H esté en mayúsculas, lo que puede ayudar a mantener la consistencia y facilitar la lectura.

def convertir_a_mayusculas_columna_h(libro_excel):

    # Recorre todas las hojas en el libro de Excel.
    # 'libro_excel.worksheets' devuelve una lista de todas las hojas en el libro.
    for hoja in libro_excel.worksheets:
        # Recorre cada fila en la columna H (columna 8), comenzando desde la segunda fila para evitar el encabezado.
        # 'iter_rows' permite iterar sobre un rango específico de celdas. 'min_col=8' y 'max_col=8' indican que solo se debe considerar la columna H.
        # 'min_row=2' asegura que se empiece desde la segunda fila para omitir el encabezado.
        for fila in hoja.iter_rows(min_col=8, max_col=8, min_row=2):
            # Obtiene la celda en la columna H para la fila actual.
            # 'fila[0]' accede a la primera celda de la fila, ya que solo estamos considerando una columna (H).
            celda = fila[0]
            
            # Verifica si la celda tiene un valor y si el valor es una cadena de texto.
            # 'celda.value' obtiene el valor de la celda y 'isinstance(celda.value, str)' asegura que el valor sea texto.
            if celda.value and isinstance(celda.value, str):
                # Cambia el texto de la celda a mayúsculas usando el método 'upper()'.
                # 'celda.value.upper()' convierte todo el texto de la celda a letras grandes.
                celda.value = celda.value.upper()



# Esta función convierte todo el texto en la columna E de un archivo Excel a letras mayúsculas.
# La columna E se selecciona dentro del archivo de Excel que se pasa como argumento (`libro_excel`).
# La función realiza los siguientes pasos:
    # 1. Accede a la hoja activa del libro de Excel.
    # 2. Recorre cada celda en la columna E.
    # 3. Toma el texto de cada celda y lo convierte a letras mayúsculas.
    # 4. Guarda el texto convertido en la misma celda.
# Esto asegura que todo el texto en la columna E esté en mayúsculas, lo que ayuda a mantener la uniformidad en los datos y facilita su lectura.

def convertir_a_mayusculas_columna_e(libro_excel):

    # Recorre todas las hojas en el libro de Excel.
    # 'libro_excel.worksheets' devuelve una lista de todas las hojas en el libro.
    for hoja in libro_excel.worksheets:
        # Recorre cada fila en la columna E (columna 5), comenzando desde la segunda fila para evitar el encabezado.
        # 'iter_rows' permite iterar sobre un rango específico de celdas. 'min_col=5' y 'max_col=5' indican que solo se debe considerar la columna E.
        # 'min_row=2' asegura que se empiece desde la segunda fila para omitir el encabezado.
        for fila in hoja.iter_rows(min_col=5, max_col=5, min_row=2):
            # Obtiene la celda en la columna E para la fila actual.
            # 'fila[0]' accede a la primera celda de la fila, ya que solo estamos considerando una columna (E).
            celda = fila[0]
            
            # Verifica si la celda tiene un valor y si el valor es una cadena de texto.
            # 'celda.value' obtiene el valor de la celda y 'isinstance(celda.value, str)' asegura que el valor sea texto.
            if celda.value and isinstance(celda.value, str):
                # Cambia el texto de la celda a mayúsculas usando el método 'upper()'.
                # 'celda.value.upper()' convierte todo el texto de la celda a letras grandes.
                celda.value = celda.value.upper()


try:
    # Define la URL del sitio web de Cinépolis en Panamá al que se desea acceder.
    # Esta URL será abierta en el navegador controlado por Selenium.
    url = "https://cinepolis.com.pa/"
    
    # Abre la página web especificada en la URL usando el navegador (driver) de Selenium.
    driver.get(url)
    
    # Crea una instancia de WebDriverWait con un tiempo de espera máximo de 10 segundos.
    # Esto permite esperar de forma dinámica hasta que los elementos de la página web estén disponibles.
    espera = WebDriverWait(driver, 10)
    
    # Llama a la función `cerrar_ventana_emergente` para cerrar cualquier ventana emergente que aparezca.
    # Esto asegura que la ventana emergente no interfiera con la selección de ciudad y cine.
    cerrar_ventana_emergente(driver, espera)
    

#1  #Panamá, David Chiriquí - Cinépolis Federal Mall
    
    # Selecciona la ciudad "Panamá, David Chiriquí" y el cine "Cinépolis Federal Mall" en la página web de Cinépolis.
    # La función `elegir_ciudad_cine` se utiliza para realizar esta selección.
    # `valor_ciudad` = "0" corresponde al atributo 'value' de la opción "Panamá, David Chiriquí" en el menú desplegable de ciudades en la página web.
    # Este valor está asociado a una opción en el elemento <select> con el id="cmbCiudades".
    # En el HTML, se verá algo así:
    # <select id="cmbCiudades" name="ctl00$cmbCiudades">
    #   <option value="Selecciona una ciudad">Selecciona una ciudad</option>
    #   <option selected="selected" value="0" clave="panama-david-chiriqui">Panamá, David Chiriquí</option>
    #   <option value="1" clave="panama-panama">Panamá, Panamá</option>
    #   <!-- Otras opciones -->
    # </select>
    #
    # `valor_cine` = "cinepolis-federal-mall" corresponde al atributo 'value' de la opción "Cinépolis Federal Mall" en el menú desplegable de cines.
    # Este valor está asociado a una opción en el elemento <select> con el id="cmbComplejos".
    # En el HTML, se verá algo así:
    # <select id="cmbComplejos" name="ctl00$cmbComplejos">
    #   <option>Selecciona un cine</option>
    #   <option value="cinepolis-federal-mall">Cinépolis Federal Mall</option>
    #   <!-- Otras opciones -->
    # </select>
    # La función selecciona la ciudad y el cine especificados en los menús desplegables y devuelve el nombre del cine seleccionado en mayúsculas.
    nombre_cine = elegir_ciudad_cine(driver, espera, "0", "cinepolis-federal-mall")


    # Imprime el nombre del cine seleccionado para verificar que la selección se realizó correctamente.
    # Este mensaje se mostrará en la consola para el usuario.
    print(f"\n\n\n\n\n\n\nNOMBRE DEL CINE SELECCIONADO (1/13): {nombre_cine}.")

    # Imprime un mensaje indicando que se está copiando los datos de la página web.
    # Se solicita al usuario que espere mientras se completa el proceso de recopilación de datos.
    print("ESTAMOS COPIANDO LOS DATOS DE LA PÁGINA WEB. POR FAVOR, ESPERE...")

    # Llama a la función `carga_completada` con un tiempo de espera de 10 segundos para asegurar que la página se haya cargado completamente.
    carga_completada(10)

    # Llama a la función `copiar_datos_a_excel` para copiar los datos de la página web a un archivo de Excel.
    # `nombre_cine` es el nombre del cine que se seleccionó, y `libro_excel` es el objeto que representa el archivo de Excel.
    copiar_datos_a_excel(nombre_cine, libro_excel)

    # Llama a la función `carga_completada` con un tiempo de espera de 5 segundos para asegurar que la página se haya cargado completamente.
    carga_completada(5)




#2  #Panamá, Panamá - Cinépolis Altaplaza

    # Selecciona la ciudad "Panamá, Panamá" y el cine "Cinépolis Altaplaza" en la página web de Cinépolis.
    # La función `elegir_ciudad_cine` se utiliza para realizar esta selección.
    # `valor_ciudad` = "1" corresponde al atributo 'value' de la opción "Panamá, Panamá" en el menú desplegable de ciudades en la página web.
    # Este valor está asociado a una opción en el elemento <select> con el id="cmbCiudades".
    # En el HTML, se verá algo así:
    # <select id="cmbCiudades" name="ctl00$cmbCiudades">
    #   <option value="Selecciona una ciudad">Selecciona una ciudad</option>
    #   <option value="0" clave="panama-david-chiriqui">Panamá, David Chiriquí</option>
    #   <option selected="selected" value="1" clave="panama-panama">Panamá, Panamá</option>
    #   <!-- Otras opciones -->
    # </select>
    #
    # `valor_cine` = "cinepolis-altaplaza" corresponde al atributo 'value' de la opción "Cinépolis Altaplaza" en el menú desplegable de cines.
    # Este valor está asociado a una opción en el elemento <select> con el id="cmbComplejos".
    # En el HTML, se verá algo así:
    # <select id="cmbComplejos" name="ctl00$cmbComplejos">
    #   <option>Selecciona un cine</option>
    #   <option value="cinepolis-altaplaza">Cinépolis Altaplaza</option>
    #   <!-- Otras opciones -->
    # </select>
    # La función selecciona la ciudad y el cine especificados en los menús desplegables y devuelve el nombre del cine seleccionado en mayúsculas.

    # Crea una instancia de WebDriverWait con un tiempo de espera máximo de 10 segundos.
    # Esto permite esperar de forma dinámica hasta que los elementos de la página web estén disponibles.
    espera = WebDriverWait(driver, 10)

    nombre_cine = elegir_ciudad_cine(driver, espera, "1", "cinepolis-altaplaza")


    # Imprime el nombre del cine seleccionado para verificar que la selección se realizó correctamente.
    # Este mensaje se mostrará en la consola para el usuario.
    print(f"\n\n\n\n\n\n\nNOMBRE DEL CINE SELECCIONADO (2/13): {nombre_cine}.")

    # Llama a la función `carga_completada` con un tiempo de espera de 10 segundos para asegurar que la página se haya cargado completamente.
    carga_completada(5)

    # Cerramos el segundo <li>, para evitar que copie datos que no corresponden al cine seleccionado
    eliminar_segundo_item_li_lista()

    # Llama a la función `carga_completada` con un tiempo de espera de 10 segundos para asegurar que la página se haya cargado completamente.
    carga_completada(5)
    
    # Imprime un mensaje indicando que se está copiando los datos de la página web.
    # Se solicita al usuario que espere mientras se completa el proceso de recopilación de datos.
    print("ESTAMOS COPIANDO LOS DATOS DE LA PÁGINA WEB. POR FAVOR, ESPERE...")
    
    
    # Llama a la función `copiar_datos_a_excel` para copiar los datos de la página web a un archivo de Excel.
    # `nombre_cine` es el nombre del cine que se seleccionó, y `libro_excel` es el objeto que representa el archivo de Excel.
    copiar_datos_a_excel(nombre_cine, libro_excel)
    
    # Llama a la función `carga_completada` con un tiempo de espera de 10 segundos para asegurar que la página se haya cargado completamente.
    carga_completada(5)







#3  #Panamá, Panamá - Cinépolis Anclas Mall  Plaza
    
    # Selecciona la ciudad "Panamá, Panamá" y el cine "Cinépolis Anclas Mall Plaza" en la página web de Cinépolis.
    # La función `elegir_ciudad_cine` se utiliza para realizar esta selección.
    # `valor_ciudad` = "1" corresponde al atributo 'value' de la opción "Panamá, Panamá" en el menú desplegable de ciudades en la página web.
    # Este valor está asociado a una opción en el elemento <select> con el id="cmbCiudades".
    # En el HTML, se verá algo así:
    # <select id="cmbCiudades" name="ctl00$cmbCiudades">
    #   <option value="Selecciona una ciudad">Selecciona una ciudad</option>
    #   <option value="0" clave="panama-david-chiriqui">Panamá, David Chiriquí</option>
    #   <option selected="selected" value="1" clave="panama-panama">Panamá, Panamá</option>
    #   <!-- Otras opciones -->
    # </select>
    #
    # `valor_cine` = "cinepolis-anclas-mall-plaza" corresponde al atributo 'value' de la opción "Cinépolis Anclas Mall Plaza" en el menú desplegable de cines.
    # Este valor está asociado a una opción en el elemento <select> con el id="cmbComplejos".
    # En el HTML, se verá algo así:
    # <select id="cmbComplejos" name="ctl00$cmbComplejos">
    #   <option>Selecciona un cine</option>
    #   <option value="cinepolis-anclas-mall-plaza">Cinépolis Anclas Mall Plaza</option>
    #   <!-- Otras opciones -->
    # </select>
    # La función selecciona la ciudad y el cine especificados en los menús desplegables y devuelve el nombre del cine seleccionado en mayúsculas.


    # Crea una instancia de WebDriverWait con un tiempo de espera máximo de 10 segundos.
    # Esto permite esperar de forma dinámica hasta que los elementos de la página web estén disponibles.
    espera = WebDriverWait(driver, 10)
    
    nombre_cine = elegir_ciudad_cine(driver, espera, "1", "cinepolis-anclas-mall-plaza")


    # Imprime el nombre del cine seleccionado para verificar que la selección se realizó correctamente.
    # Este mensaje se mostrará en la consola para el usuario.
    print(f"\n\n\n\n\n\n\nNOMBRE DEL CINE SELECCIONADO (3/13): {nombre_cine}.")

    # Imprime un mensaje indicando que se está copiando los datos de la página web.
    # Se solicita al usuario que espere mientras se completa el proceso de recopilación de datos.
    print("ESTAMOS COPIANDO LOS DATOS DE LA PÁGINA WEB. POR FAVOR, ESPERE...")

    # Llama a la función `carga_completada` con un tiempo de espera de 10 segundos para asegurar que la página se haya cargado completamente.
    carga_completada(10)

    # Llama a la función `copiar_datos_a_excel` para copiar los datos de la página web a un archivo de Excel.
    # `nombre_cine` es el nombre del cine que se seleccionó, y `libro_excel` es el objeto que representa el archivo de Excel.
    copiar_datos_a_excel(nombre_cine, libro_excel)

    # Llama a la función `carga_completada` con un tiempo de espera de 5 segundos para asegurar que la página se haya cargado completamente.
    carga_completada(5)


#4  #Panamá, Panamá - Cinépolis Andes Mall

    # Selecciona la ciudad "Panamá, Panamá" y el cine "Cinépolis Andes Mall" en la página web de Cinépolis.
    # La función `elegir_ciudad_cine` se utiliza para realizar esta selección.
    # `valor_ciudad` = "1" corresponde al atributo 'value' de la opción "Panamá, Panamá" en el menú desplegable de ciudades en la página web.
    # Este valor está asociado a una opción en el elemento <select> con el id="cmbCiudades".
    # En el HTML, se verá algo así:
    # <select id="cmbCiudades" name="ctl00$cmbCiudades">
    #   <option value="Selecciona una ciudad">Selecciona una ciudad</option>
    #   <option value="0" clave="panama-david-chiriqui">Panamá, David Chiriquí</option>
    #   <option selected="selected" value="1" clave="panama-panama">Panamá, Panamá</option>
    #   <!-- Otras opciones -->
    # </select>
    #
    # `valor_cine` = "cinepolis-andes-mall" corresponde al atributo 'value' de la opción "Cinépolis Andes Mall" en el menú desplegable de cines.
    # Este valor está asociado a una opción en el elemento <select> con el id="cmbComplejos".
    # En el HTML, se verá algo así:
    # <select id="cmbComplejos" name="ctl00$cmbComplejos">
    #   <option>Selecciona un cine</option>
    #   <option value="cinepolis-andes-mall">Cinépolis Andes Mall</option>
    #   <!-- Otras opciones -->
    # </select>
    # La función selecciona la ciudad y el cine especificados en los menús desplegables y devuelve el nombre del cine seleccionado en mayúsculas.

    # Crea una instancia de WebDriverWait con un tiempo de espera máximo de 10 segundos.
    # Esto permite esperar de forma dinámica hasta que los elementos de la página web estén disponibles.
    espera = WebDriverWait(driver, 10)
    
    nombre_cine = elegir_ciudad_cine(driver, espera, "1", "cinepolis-andes-mall")


    # Imprime el nombre del cine seleccionado para verificar que la selección se realizó correctamente.
    # Este mensaje se mostrará en la consola para el usuario.
    print(f"\n\n\n\n\n\n\nNOMBRE DEL CINE SELECCIONADO (4/13): {nombre_cine}.")

    # Imprime un mensaje indicando que se está copiando los datos de la página web.
    # Se solicita al usuario que espere mientras se completa el proceso de recopilación de datos.
    print("ESTAMOS COPIANDO LOS DATOS DE LA PÁGINA WEB. POR FAVOR, ESPERE...")

    # Llama a la función `carga_completada` con un tiempo de espera de 10 segundos para asegurar que la página se haya cargado completamente.
    carga_completada(10)

    # Llama a la función `copiar_datos_a_excel` para copiar los datos de la página web a un archivo de Excel.
    # `nombre_cine` es el nombre del cine que se seleccionó, y `libro_excel` es el objeto que representa el archivo de Excel.
    copiar_datos_a_excel(nombre_cine, libro_excel)

    # Llama a la función `carga_completada` con un tiempo de espera de 5 segundos para asegurar que la página se haya cargado completamente.
    carga_completada(5)






#5  #Panamá, Panamá - Cinépolis El Dorado
    
    # Selecciona la ciudad "Panamá, Panamá" y el cine "Cinépolis El Dorado" en la página web de Cinépolis.
    # La función `elegir_ciudad_cine` se utiliza para realizar esta selección.
    # `valor_ciudad` = "1" corresponde al atributo 'value' de la opción "Panamá, Panamá" en el menú desplegable de ciudades en la página web.
    # Este valor está asociado a una opción en el elemento <select> con el id="cmbCiudades".
    # En el HTML, se verá algo así:
    # <select id="cmbCiudades" name="ctl00$cmbCiudades">
    #   <option value="Selecciona una ciudad">Selecciona una ciudad</option>
    #   <option value="0" clave="panama-david-chiriqui">Panamá, David Chiriquí</option>
    #   <option selected="selected" value="1" clave="panama-panama">Panamá, Panamá</option>
    #   <!-- Otras opciones -->
    # </select>
    #
    # `valor_cine` = "cinepolis-el-dorado" corresponde al atributo 'value' de la opción "Cinépolis El Dorado" en el menú desplegable de cines.
    # Este valor está asociado a una opción en el elemento <select> con el id="cmbComplejos".
    # En el HTML, se verá algo así:
    # <select id="cmbComplejos" name="ctl00$cmbComplejos">
    #   <option>Selecciona un cine</option>
    #   <option value="cinepolis-el-dorado">Cinépolis El Dorado</option>
    #   <!-- Otras opciones -->
    # </select>
    # La función selecciona la ciudad y el cine especificados en los menús desplegables y devuelve el nombre del cine seleccionado en mayúsculas.

    # Crea una instancia de WebDriverWait con un tiempo de espera máximo de 10 segundos.
    # Esto permite esperar de forma dinámica hasta que los elementos de la página web estén disponibles.
    espera = WebDriverWait(driver, 10)
    
    nombre_cine = elegir_ciudad_cine(driver, espera, "1", "cinepolis-el-dorado")


    # Imprime el nombre del cine seleccionado para verificar que la selección se realizó correctamente.
    # Este mensaje se mostrará en la consola para el usuario.
    print(f"\n\n\n\n\n\n\nNOMBRE DEL CINE SELECCIONADO (5/13): {nombre_cine}.")

    # Imprime un mensaje indicando que se está copiando los datos de la página web.
    # Se solicita al usuario que espere mientras se completa el proceso de recopilación de datos.
    print("ESTAMOS COPIANDO LOS DATOS DE LA PÁGINA WEB. POR FAVOR, ESPERE...")

    # Llama a la función `carga_completada` con un tiempo de espera de 10 segundos para asegurar que la página se haya cargado completamente.
    carga_completada(10)

    # Llama a la función `copiar_datos_a_excel` para copiar los datos de la página web a un archivo de Excel.
    # `nombre_cine` es el nombre del cine que se seleccionó, y `libro_excel` es el objeto que representa el archivo de Excel.
    copiar_datos_a_excel(nombre_cine, libro_excel)

    # Llama a la función `carga_completada` con un tiempo de espera de 5 segundos para asegurar que la página se haya cargado completamente.
    carga_completada(5)



#6  #Panamá, Panamá - Cinépolis Metro Mall
    
    # Selecciona la ciudad "Panamá, Panamá" y el cine "Cinépolis Metro Mall" en la página web de Cinépolis.
    # La función `elegir_ciudad_cine` se utiliza para realizar esta selección.
    # `valor_ciudad` = "1" corresponde al atributo 'value' de la opción "Panamá, Panamá" en el menú desplegable de ciudades en la página web.
    # Este valor está asociado a una opción en el elemento <select> con el id="cmbCiudades".
    # En el HTML, se verá algo así:
    # <select id="cmbCiudades" name="ctl00$cmbCiudades">
    #   <option value="Selecciona una ciudad">Selecciona una ciudad</option>
    #   <option value="0" clave="panama-david-chiriqui">Panamá, David Chiriquí</option>
    #   <option selected="selected" value="1" clave="panama-panama">Panamá, Panamá</option>
    #   <!-- Otras opciones -->
    # </select>
    #
    # `valor_cine` = "cinepolis-metro-mall" corresponde al atributo 'value' de la opción "Cinépolis Metro Mall" en el menú desplegable de cines.
    # Este valor está asociado a una opción en el elemento <select> con el id="cmbComplejos".
    # En el HTML, se verá algo así:
    # <select id="cmbComplejos" name="ctl00$cmbComplejos">
    #   <option>Selecciona un cine</option>
    #   <option value="cinepolis-metro-mall">Cinépolis Metro Mall</option>
    #   <!-- Otras opciones -->
    # </select>
    # La función selecciona la ciudad y el cine especificados en los menús desplegables y devuelve el nombre del cine seleccionado en mayúsculas.

    # Crea una instancia de WebDriverWait con un tiempo de espera máximo de 10 segundos.
    # Esto permite esperar de forma dinámica hasta que los elementos de la página web estén disponibles.
    espera = WebDriverWait(driver, 10)
    
    nombre_cine = elegir_ciudad_cine(driver, espera, "1", "cinepolis-metro-mall")


    # Imprime el nombre del cine seleccionado para verificar que la selección se realizó correctamente.
    # Este mensaje se mostrará en la consola para el usuario.
    print(f"\n\n\n\n\n\n\nNOMBRE DEL CINE SELECCIONADO (6/13): {nombre_cine}.")

    # Imprime un mensaje indicando que se está copiando los datos de la página web.
    # Se solicita al usuario que espere mientras se completa el proceso de recopilación de datos.
    print("ESTAMOS COPIANDO LOS DATOS DE LA PÁGINA WEB. POR FAVOR, ESPERE...")

    # Llama a la función `carga_completada` con un tiempo de espera de 10 segundos para asegurar que la página se haya cargado completamente.
    carga_completada(10)

    # Llama a la función `copiar_datos_a_excel` para copiar los datos de la página web a un archivo de Excel.
    # `nombre_cine` es el nombre del cine que se seleccionó, y `libro_excel` es el objeto que representa el archivo de Excel.
    copiar_datos_a_excel(nombre_cine, libro_excel)

    # Llama a la función `carga_completada` con un tiempo de espera de 5 segundos para asegurar que la página se haya cargado completamente.
    carga_completada(5)



#7  #Panamá, Panamá - Cinépolis Multiplaza Pacific
    
    # Selecciona la ciudad "Panamá, Panamá" y el cine "Cinépolis Multiplaza Pacific" en la página web de Cinépolis.
    # La función `elegir_ciudad_cine` se utiliza para realizar esta selección.
    # `valor_ciudad` = "1" corresponde al atributo 'value' de la opción "Panamá, Panamá" en el menú desplegable de ciudades en la página web.
    # Este valor está asociado a una opción en el elemento <select> con el id="cmbCiudades".
    # En el HTML, se verá algo así:
    # <select id="cmbCiudades" name="ctl00$cmbCiudades">
    #   <option value="Selecciona una ciudad">Selecciona una ciudad</option>
    #   <option value="0" clave="panama-david-chiriqui">Panamá, David Chiriquí</option>
    #   <option selected="selected" value="1" clave="panama-panama">Panamá, Panamá</option>
    #   <!-- Otras opciones -->
    # </select>
    #
    # `valor_cine` = "cinepolis-multiplaza-pacific" corresponde al atributo 'value' de la opción "Cinépolis Multiplaza Pacific" en el menú desplegable de cines.
    # Este valor está asociado a una opción en el elemento <select> con el id="cmbComplejos".
    # En el HTML, se verá algo así:
    # <select id="cmbComplejos" name="ctl00$cmbComplejos">
    #   <option>Selecciona un cine</option>
    #   <option value="cinepolis-multiplaza-pacific">Cinépolis Multiplaza Pacific</option>
    #   <!-- Otras opciones -->
    # </select>
    # La función selecciona la ciudad y el cine especificados en los menús desplegables y devuelve el nombre del cine seleccionado en mayúsculas.

    # Crea una instancia de WebDriverWait con un tiempo de espera máximo de 10 segundos.
    # Esto permite esperar de forma dinámica hasta que los elementos de la página web estén disponibles.
    espera = WebDriverWait(driver, 10)
    
    nombre_cine = elegir_ciudad_cine(driver, espera, "1", "cinepolis-multiplaza-pacific")


    # Imprime el nombre del cine seleccionado para verificar que la selección se realizó correctamente.
    # Este mensaje se mostrará en la consola para el usuario.
    print(f"\n\n\n\n\n\n\nNOMBRE DEL CINE SELECCIONADO (7/13): {nombre_cine}.")

    # Llama a la función `carga_completada` con un tiempo de espera de 10 segundos para asegurar que la página se haya cargado completamente.
    carga_completada(5)

    # Cerramos el segundo <li>, para evitar que copie datos que no corresponden al cine seleccionado
    eliminar_segundo_item_li_lista()

    # Llama a la función `carga_completada` con un tiempo de espera de 10 segundos para asegurar que la página se haya cargado completamente.
    carga_completada(5)

    # Imprime un mensaje indicando que se está copiando los datos de la página web.
    # Se solicita al usuario que espere mientras se completa el proceso de recopilación de datos.
    print("ESTAMOS COPIANDO LOS DATOS DE LA PÁGINA WEB. POR FAVOR, ESPERE...")

    # Llama a la función `carga_completada` con un tiempo de espera de 10 segundos para asegurar que la página se haya cargado completamente.
    carga_completada(10)

    # Llama a la función `copiar_datos_a_excel` para copiar los datos de la página web a un archivo de Excel.
    # `nombre_cine` es el nombre del cine que se seleccionó, y `libro_excel` es el objeto que representa el archivo de Excel.
    copiar_datos_a_excel(nombre_cine, libro_excel)

    # Llama a la función `carga_completada` con un tiempo de espera de 5 segundos para asegurar que la página se haya cargado completamente.
    carga_completada(5)




#8  #Panamá, Panamá - Cinépolis Town Center

    # Selecciona la ciudad "Panamá, Panamá" y el cine "Cinépolis Town Center" en la página web de Cinépolis.
    # La función `elegir_ciudad_cine` se utiliza para realizar esta selección.
    # `valor_ciudad` = "1" corresponde al atributo 'value' de la opción "Panamá, Panamá" en el menú desplegable de ciudades en la página web.
    # Este valor está asociado a una opción en el elemento <select> con el id="cmbCiudades".
    # En el HTML, se verá algo así:
    # <select id="cmbCiudades" name="ctl00$cmbCiudades">
    #   <option value="Selecciona una ciudad">Selecciona una ciudad</option>
    #   <option value="0" clave="panama-david-chiriqui">Panamá, David Chiriquí</option>
    #   <option selected="selected" value="1" clave="panama-panama">Panamá, Panamá</option>
    #   <!-- Otras opciones -->
    # </select>
    #
    # `valor_cine` = "cinepolis-town-center" corresponde al atributo 'value' de la opción "Cinépolis Town Center" en el menú desplegable de cines.
    # Este valor está asociado a una opción en el elemento <select> con el id="cmbComplejos".
    # En el HTML, se verá algo así:
    # <select id="cmbComplejos" name="ctl00$cmbComplejos">
    #   <option>Selecciona un cine</option>
    #   <option value="cinepolis-town-center">Cinépolis Town Center</option>
    #   <!-- Otras opciones -->
    # </select>
    # La función selecciona la ciudad y el cine especificados en los menús desplegables y devuelve el nombre del cine seleccionado en mayúsculas.

    # Crea una instancia de WebDriverWait con un tiempo de espera máximo de 10 segundos.
    # Esto permite esperar de forma dinámica hasta que los elementos de la página web estén disponibles.
    espera = WebDriverWait(driver, 10)
    
    nombre_cine = elegir_ciudad_cine(driver, espera, "1", "cinepolis-town-center")


    # Imprime el nombre del cine seleccionado para verificar que la selección se realizó correctamente.
    # Este mensaje se mostrará en la consola para el usuario.
    print(f"\n\n\n\n\n\n\nNOMBRE DEL CINE SELECCIONADO (8/13): {nombre_cine}.")

    # Imprime un mensaje indicando que se está copiando los datos de la página web.
    # Se solicita al usuario que espere mientras se completa el proceso de recopilación de datos.
    print("ESTAMOS COPIANDO LOS DATOS DE LA PÁGINA WEB. POR FAVOR, ESPERE...")

    # Llama a la función `carga_completada` con un tiempo de espera de 10 segundos para asegurar que la página se haya cargado completamente.
    carga_completada(10)

    # Llama a la función `copiar_datos_a_excel` para copiar los datos de la página web a un archivo de Excel.
    # `nombre_cine` es el nombre del cine que se seleccionó, y `libro_excel` es el objeto que representa el archivo de Excel.
    copiar_datos_a_excel(nombre_cine, libro_excel)

    # Llama a la función `carga_completada` con un tiempo de espera de 5 segundos para asegurar que la página se haya cargado completamente.
    carga_completada(5)



#9  #Panamá, Panamá - Cinépolis VIP Altaplaza

    # Selecciona la ciudad "Panamá, Panamá" y el cine "Cinépolis VIP Altaplaza" en la página web de Cinépolis.
    # La función `elegir_ciudad_cine` se utiliza para realizar esta selección.
    # `valor_ciudad` = "1" corresponde al atributo 'value' de la opción "Panamá, Panamá" en el menú desplegable de ciudades en la página web.
    # Este valor está asociado a una opción en el elemento <select> con el id="cmbCiudades".
    # En el HTML, se verá algo así:
    # <select id="cmbCiudades" name="ctl00$cmbCiudades">
    #   <option value="Selecciona una ciudad">Selecciona una ciudad</option>
    #   <option value="0" clave="panama-david-chiriqui">Panamá, David Chiriquí</option>
    #   <option selected="selected" value="1" clave="panama-panama">Panamá, Panamá</option>
    #   <!-- Otras opciones -->
    # </select>
    #
    # `valor_cine` = "cinepolis-vip-altaplaza" corresponde al atributo 'value' de la opción "Cinépolis VIP Altaplaza" en el menú desplegable de cines.
    # Este valor está asociado a una opción en el elemento <select> con el id="cmbComplejos".
    # En el HTML, se verá algo así:
    # <select id="cmbComplejos" name="ctl00$cmbComplejos">
    #   <option>Selecciona un cine</option>
    #   <option value="cinepolis-vip-altaplaza">Cinépolis VIP Altaplaza</option>
    #   <!-- Otras opciones -->
    # </select>
    # La función selecciona la ciudad y el cine especificados en los menús desplegables y devuelve el nombre del cine seleccionado en mayúsculas.

    # Crea una instancia de WebDriverWait con un tiempo de espera máximo de 10 segundos.
    # Esto permite esperar de forma dinámica hasta que los elementos de la página web estén disponibles.
    espera = WebDriverWait(driver, 10)
    
    nombre_cine = elegir_ciudad_cine(driver, espera, "1", "cinepolis-vip-altaplaza")


    # Imprime el nombre del cine seleccionado para verificar que la selección se realizó correctamente.
    # Este mensaje se mostrará en la consola para el usuario.
    print(f"\n\n\n\n\n\n\nNOMBRE DEL CINE SELECCIONADO (9/13): {nombre_cine}.")

    # Imprime un mensaje indicando que se está copiando los datos de la página web.
    # Se solicita al usuario que espere mientras se completa el proceso de recopilación de datos.
    print("ESTAMOS COPIANDO LOS DATOS DE LA PÁGINA WEB. POR FAVOR, ESPERE...")

    # Llama a la función `carga_completada` con un tiempo de espera de 10 segundos para asegurar que la página se haya cargado completamente.
    carga_completada(10)

    # Llama a la función `copiar_datos_a_excel` para copiar los datos de la página web a un archivo de Excel.
    # `nombre_cine` es el nombre del cine que se seleccionó, y `libro_excel` es el objeto que representa el archivo de Excel.
    copiar_datos_a_excel(nombre_cine, libro_excel)

    # Llama a la función `carga_completada` con un tiempo de espera de 5 segundos para asegurar que la página se haya cargado completamente.
    carga_completada(5)




#10  #Panamá, Panamá - Cinépolis VIP Multiplaza Pacific

    # Selecciona la ciudad "Panamá, Panamá" y el cine "Cinépolis VIP Multiplaza Pacific" en la página web de Cinépolis.
    # La función `elegir_ciudad_cine` se utiliza para realizar esta selección.
    # `valor_ciudad` = "1" corresponde al atributo 'value' de la opción "Panamá, Panamá" en el menú desplegable de ciudades en la página web.
    # Este valor está asociado a una opción en el elemento <select> con el id="cmbCiudades".
    # En el HTML, se verá algo así:
    # <select id="cmbCiudades" name="ctl00$cmbCiudades">
    #   <option value="Selecciona una ciudad">Selecciona una ciudad</option>
    #   <option value="0" clave="panama-david-chiriqui">Panamá, David Chiriquí</option>
    #   <option selected="selected" value="1" clave="panama-panama">Panamá, Panamá</option>
    #   <!-- Otras opciones -->
    # </select>
    #
    # `valor_cine` = "cinepolis-vip-multiplaza-pacific" corresponde al atributo 'value' de la opción "Cinépolis VIP Multiplaza Pacific" en el menú desplegable de cines.
    # Este valor está asociado a una opción en el elemento <select> con el id="cmbComplejos".
    # En el HTML, se verá algo así:
    # <select id="cmbComplejos" name="ctl00$cmbComplejos">
    #   <option>Selecciona un cine</option>
    #   <option value="cinepolis-vip-multiplaza-pacific">Cinépolis VIP Multiplaza Pacific</option>
    #   <!-- Otras opciones -->
    # </select>
    # La función selecciona la ciudad y el cine especificados en los menús desplegables y devuelve el nombre del cine seleccionado en mayúsculas.

    # Crea una instancia de WebDriverWait con un tiempo de espera máximo de 10 segundos.
    # Esto permite esperar de forma dinámica hasta que los elementos de la página web estén disponibles.
    espera = WebDriverWait(driver, 10)
    
    nombre_cine = elegir_ciudad_cine(driver, espera, "1", "cinepolis-vip-multiplaza-pacific")


    # Imprime el nombre del cine seleccionado para verificar que la selección se realizó correctamente.
    # Este mensaje se mostrará en la consola para el usuario.
    print(f"\n\n\n\n\n\n\nNOMBRE DEL CINE SELECCIONADO (10/13): {nombre_cine}.")

    # Llama a la función `carga_completada` con un tiempo de espera de 10 segundos para asegurar que la página se haya cargado completamente.
    carga_completada(5)

    # Cerramos el primer <li>, para evitar que copie datos que no corresponden al cine seleccionado
    eliminar_primer_item_li_lista()

    # Llama a la función `carga_completada` con un tiempo de espera de 10 segundos para asegurar que la página se haya cargado completamente.
    carga_completada(5)

    # Imprime un mensaje indicando que se está copiando los datos de la página web.
    # Se solicita al usuario que espere mientras se completa el proceso de recopilación de datos.
    print("ESTAMOS COPIANDO LOS DATOS DE LA PÁGINA WEB. POR FAVOR, ESPERE...")

    # Llama a la función `carga_completada` con un tiempo de espera de 10 segundos para asegurar que la página se haya cargado completamente.
    carga_completada(10)

    # Llama a la función `copiar_datos_a_excel` para copiar los datos de la página web a un archivo de Excel.
    # `nombre_cine` es el nombre del cine que se seleccionó, y `libro_excel` es el objeto que representa el archivo de Excel.
    copiar_datos_a_excel(nombre_cine, libro_excel)

    # Llama a la función `carga_completada` con un tiempo de espera de 5 segundos para asegurar que la página se haya cargado completamente.
    carga_completada(5)




#11  #Panamá, Panamá - Cinépolis VIP Soho

    # Selecciona la ciudad "Panamá, Panamá" y el cine "Cinépolis VIP Soho" en la página web de Cinépolis.
    # La función `elegir_ciudad_cine` se utiliza para realizar esta selección.
    # `valor_ciudad` = "1" corresponde al atributo 'value' de la opción "Panamá, Panamá" en el menú desplegable de ciudades en la página web.
    # Este valor está asociado a una opción en el elemento <select> con el id="cmbCiudades".
    # En el HTML, se verá algo así:
    # <select id="cmbCiudades" name="ctl00$cmbCiudades">
    #   <option value="Selecciona una ciudad">Selecciona una ciudad</option>
    #   <option value="0" clave="panama-david-chiriqui">Panamá, David Chiriquí</option>
    #   <option selected="selected" value="1" clave="panama-panama">Panamá, Panamá</option>
    #   <!-- Otras opciones -->
    # </select>
    #
    # `valor_cine` = "cinepolis-vip-soho" corresponde al atributo 'value' de la opción "Cinépolis VIP Soho" en el menú desplegable de cines.
    # Este valor está asociado a una opción en el elemento <select> con el id="cmbComplejos".
    # En el HTML, se verá algo así:
    # <select id="cmbComplejos" name="ctl00$cmbComplejos">
    #   <option>Selecciona un cine</option>
    #   <option value="cinepolis-vip-soho">Cinépolis VIP Soho</option>
    #   <!-- Otras opciones -->
    # </select>
    # La función selecciona la ciudad y el cine especificados en los menús desplegables y devuelve el nombre del cine seleccionado en mayúsculas.

    # Crea una instancia de WebDriverWait con un tiempo de espera máximo de 10 segundos.
    # Esto permite esperar de forma dinámica hasta que los elementos de la página web estén disponibles.
    espera = WebDriverWait(driver, 10)
    
    nombre_cine = elegir_ciudad_cine(driver, espera, "1", "cinepolis-vip-soho")


    # Imprime el nombre del cine seleccionado para verificar que la selección se realizó correctamente.
    # Este mensaje se mostrará en la consola para el usuario.
    print(f"\n\n\n\n\n\n\nNOMBRE DEL CINE SELECCIONADO (11/13): {nombre_cine}.")

    # Imprime un mensaje indicando que se está copiando los datos de la página web.
    # Se solicita al usuario que espere mientras se completa el proceso de recopilación de datos.
    print("ESTAMOS COPIANDO LOS DATOS DE LA PÁGINA WEB. POR FAVOR, ESPERE...")

    # Llama a la función `carga_completada` con un tiempo de espera de 10 segundos para asegurar que la página se haya cargado completamente.
    carga_completada(10)

    # Llama a la función `copiar_datos_a_excel` para copiar los datos de la página web a un archivo de Excel.
    # `nombre_cine` es el nombre del cine que se seleccionó, y `libro_excel` es el objeto que representa el archivo de Excel.
    copiar_datos_a_excel(nombre_cine, libro_excel)

    # Llama a la función `carga_completada` con un tiempo de espera de 5 segundos para asegurar que la página se haya cargado completamente.
    carga_completada(5)


#12  #Panamá, Panamá - Cinépolis VIP Town Center

    # Selecciona la ciudad "Panamá, Panamá" y el cine "Cinépolis VIP Town Center" en la página web de Cinépolis.
    # La función `elegir_ciudad_cine` se utiliza para realizar esta selección.
    # `valor_ciudad` = "1" corresponde al atributo 'value' de la opción "Panamá, Panamá" en el menú desplegable de ciudades en la página web.
    # Este valor está asociado a una opción en el elemento <select> con el id="cmbCiudades".
    # En el HTML, se verá algo así:
    # <select id="cmbCiudades" name="ctl00$cmbCiudades">
    #   <option value="Selecciona una ciudad">Selecciona una ciudad</option>
    #   <option value="0" clave="panama-david-chiriqui">Panamá, David Chiriquí</option>
    #   <option selected="selected" value="1" clave="panama-panama">Panamá, Panamá</option>
    #   <!-- Otras opciones -->
    # </select>
    #
    # `valor_cine` = "cinepolis-vip-town-center" corresponde al atributo 'value' de la opción "Cinépolis VIP Town Center" en el menú desplegable de cines.
    # Este valor está asociado a una opción en el elemento <select> con el id="cmbComplejos".
    # En el HTML, se verá algo así:
    # <select id="cmbComplejos" name="ctl00$cmbComplejos">
    #   <option>Selecciona un cine</option>
    #   <option value="cinepolis-vip-town-center">Cinépolis VIP Town Center</option>
    #   <!-- Otras opciones -->
    # </select>
    # La función selecciona la ciudad y el cine especificados en los menús desplegables y devuelve el nombre del cine seleccionado en mayúsculas.

    # Crea una instancia de WebDriverWait con un tiempo de espera máximo de 10 segundos.
    # Esto permite esperar de forma dinámica hasta que los elementos de la página web estén disponibles.
    espera = WebDriverWait(driver, 10)
    
    nombre_cine = elegir_ciudad_cine(driver, espera, "1", "cinepolis-vip-town-center")


    # Imprime el nombre del cine seleccionado para verificar que la selección se realizó correctamente.
    # Este mensaje se mostrará en la consola para el usuario.
    print(f"\n\n\n\n\n\n\nNOMBRE DEL CINE SELECCIONADO (12/13): {nombre_cine}.")

    # Imprime un mensaje indicando que se está copiando los datos de la página web.
    # Se solicita al usuario que espere mientras se completa el proceso de recopilación de datos.
    print("ESTAMOS COPIANDO LOS DATOS DE LA PÁGINA WEB. POR FAVOR, ESPERE...")

    # Llama a la función `carga_completada` con un tiempo de espera de 10 segundos para asegurar que la página se haya cargado completamente.
    carga_completada(10)

    # Llama a la función `copiar_datos_a_excel` para copiar los datos de la página web a un archivo de Excel.
    # `nombre_cine` es el nombre del cine que se seleccionó, y `libro_excel` es el objeto que representa el archivo de Excel.
    copiar_datos_a_excel(nombre_cine, libro_excel)

    # Llama a la función `carga_completada` con un tiempo de espera de 5 segundos para asegurar que la página se haya cargado completamente.
    carga_completada(5)




#13  #Panamá, Panamá - Cinépolis Westland Mall

    # Selecciona la ciudad "Panamá, Panamá" y el cine "Cinépolis Westland Mall" en la página web de Cinépolis.
    # La función `elegir_ciudad_cine` se utiliza para realizar esta selección.
    # `valor_ciudad` = "1" corresponde al atributo 'value' de la opción "Panamá, Panamá" en el menú desplegable de ciudades en la página web.
    # Este valor está asociado a una opción en el elemento <select> con el id="cmbCiudades".
    # En el HTML, se verá algo así:
    # <select id="cmbCiudades" name="ctl00$cmbCiudades">
    #   <option value="Selecciona una ciudad">Selecciona una ciudad</option>
    #   <option value="0" clave="panama-david-chiriqui">Panamá, David Chiriquí</option>
    #   <option selected="selected" value="1" clave="panama-panama">Panamá, Panamá</option>
    #   <!-- Otras opciones -->
    # </select>
    #
    # `valor_cine` = "cinepolis-westland-mall" corresponde al atributo 'value' de la opción "Cinépolis Westland Mall" en el menú desplegable de cines.
    # Este valor está asociado a una opción en el elemento <select> con el id="cmbComplejos".
    # En el HTML, se verá algo así:
    # <select id="cmbComplejos" name="ctl00$cmbComplejos">
    #   <option>Selecciona un cine</option>
    #   <option value="cinepolis-westland-mall">Cinépolis Westland Mall</option>
    #   <!-- Otras opciones -->
    # </select>
    # La función selecciona la ciudad y el cine especificados en los menús desplegables y devuelve el nombre del cine seleccionado en mayúsculas.

    # Crea una instancia de WebDriverWait con un tiempo de espera máximo de 10 segundos.
    # Esto permite esperar de forma dinámica hasta que los elementos de la página web estén disponibles.
    espera = WebDriverWait(driver, 10)
    
    nombre_cine = elegir_ciudad_cine(driver, espera, "1", "cinepolis-westland-mall")


    # Imprime el nombre del cine seleccionado para verificar que la selección se realizó correctamente.
    # Este mensaje se mostrará en la consola para el usuario.
    print(f"\n\n\n\n\n\n\nNOMBRE DEL CINE SELECCIONADO (13/13): {nombre_cine}.")

    # Imprime un mensaje indicando que se está copiando los datos de la página web.
    # Se solicita al usuario que espere mientras se completa el proceso de recopilación de datos.
    print("ESTAMOS COPIANDO LOS DATOS DE LA PÁGINA WEB. POR FAVOR, ESPERE...")

    # Llama a la función `carga_completada` con un tiempo de espera de 10 segundos para asegurar que la página se haya cargado completamente.
    carga_completada(10)

    # Llama a la función `copiar_datos_a_excel` para copiar los datos de la página web a un archivo de Excel.
    # `nombre_cine` es el nombre del cine que se seleccionó, y `libro_excel` es el objeto que representa el archivo de Excel.
    copiar_datos_a_excel(nombre_cine, libro_excel)

    # Llama a la función `carga_completada` con un tiempo de espera de 5 segundos para asegurar que la página se haya cargado completamente.
    carga_completada(5)




    

# Esta línea de código maneja errores que puedan ocurrir durante la ejecución del programa.
# Si ocurre algún problema (como un error inesperado), se captura el error y se muestra un mensaje.
except Exception as e:
    # Imprime un mensaje en la pantalla que indica que ocurrió un error durante la ejecución.
    # El mensaje incluye detalles sobre el error, que se almacena en la variable 'e'.
    print(f"Error en la ejecución: {e}")



# El bloque 'finally' se ejecuta siempre, sin importar si hubo un error o no.
# Esto significa que, sin importar si el código anterior terminó correctamente o si hubo un error,
# el código dentro del bloque 'finally' siempre se ejecutará al final.
finally:
    # Esta línea crea el nombre del archivo Excel que se guardará.
    # El nombre del archivo tendrá el formato "CINEPOLIS-PANAMÁ-YYYY-MM-DD.xlsx", donde YYYY-MM-DD es la fecha actual.
    # Por ejemplo, si hoy es el 12 de agosto de 2024, el nombre del archivo será "CINEPOLIS-PANAMÁ-2024-08-12.xlsx".
    nombre_archivo = f"CINEPOLIS-PANAMÁ-{datetime.now().strftime('%Y-%m-%d')}.xlsx"

    # Esta línea determina la ubicación completa del archivo que se va a guardar.
    # 'os.getcwd()' obtiene la ruta del directorio actual donde se está ejecutando el programa.
    # 'os.path.join()' combina esta ruta con el nombre del archivo para crear la ruta completa del archivo.
    # Así, 'ruta_archivo' es la dirección completa donde el archivo Excel se guardará en el sistema.
    ruta_archivo = os.path.join(os.getcwd(), nombre_archivo)


    # Muestra un mensaje en la pantalla indicando que se está guardando la información en múltiples hojas del archivo Excel.
    print("\n\n\n\n\n\nGUARDANDO LA INFORMACIÓN EN MULTIPLES HOJAS DEL ARCHIVO EXCEL...")

    # Llama a una función para eliminar iconos o símbolos en la columna H del archivo Excel.
    eliminar_icono_de_columna_h(libro_excel)

    # Hace que el programa espere 2 segundos para asegurarse de que los cambios en la columna H se hayan aplicado correctamente.
    time.sleep(2)

    # Llama a una función para eliminar palabras repetidas en la columna H del archivo Excel.
    eliminar_palabras_repetidas_columna_h(libro_excel)

    # Hace que el programa espere 2 segundos para asegurarse de que los cambios se hayan aplicado antes de continuar.
    time.sleep(2)

    # Llama a una función para convertir todo el texto en la columna E del archivo Excel a mayúsculas.
    convertir_a_mayusculas_columna_e(libro_excel)

    # Hace que el programa espere 2 segundos para asegurarse de que los cambios se hayan aplicado antes de continuar.
    time.sleep(2)

    # Llama a una función para convertir todo el texto en la columna H del archivo Excel a mayúsculas.
    convertir_a_mayusculas_columna_h(libro_excel)

    # Se realiza una pausa de cinco segundos para asegurar que el proceso se complete correctamente y minimizar el riesgo de errores.
    carga_completada(5)

    # Guarda el archivo de Excel con todos los datos y cambios que hemos hecho.
    # 'ruta_archivo' es la ubicación y nombre del archivo donde se guardará.
    libro_excel.save(ruta_archivo)

    # Imprime un mensaje en la consola que indica que el archivo de Excel se ha guardado correctamente.
    # 'nombre_archivo' muestra el nombre del archivo donde se guardaron las hojas.
    print(f"¡LAS HOJAS EXCEL SE HAN GUARDADO EN EL ARCHIVO EXCEL: {nombre_archivo}.!")

    # Pausa la ejecución del programa y espera a que el usuario presione ENTER antes de continuar.
    # Esto permite al usuario leer el mensaje anterior antes de cerrar el navegador.
    input("\n\n Presiona ENTER para cerrar el navegador y salir de la consola Python.")

    # Cierra el navegador web y termina la sesión del controlador (driver).
    # Esto es necesario para liberar los recursos del sistema una vez que hemos terminado.
    driver.quit()

