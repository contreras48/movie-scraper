# Importamos el módulo 'os', que nos permite interactuar con el sistema operativo, como manejar archivos y directorios.
import os

# Importamos el módulo 'time', que nos permite trabajar con el tiempo, por ejemplo, hacer pausas en la ejecución del programa.
import time

# Importamos el módulo 'sys', que nos permite interactuar con el sistema de entrada y salida del programa, como leer argumentos de la línea de comandos.
import sys

# Importamos la clase 'datetime' del módulo 'datetime', que nos permite trabajar con fechas y horas.
from datetime import datetime

# Importamos el módulo 'openpyxl', que nos permite leer y escribir archivos de Excel.
import openpyxl

# Importamos las clases y funciones necesarias de 'selenium', que nos permiten automatizar la interacción con navegadores web.
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.edge.service import Service
from selenium.webdriver.edge.options import Options

# Importamos la función 'urlparse' del módulo 'urllib.parse', que nos permite analizar y descomponer URLs (direcciones web).
from urllib.parse import urlparse


# Creamos un objeto para configurar las opciones del navegador Microsoft Edge.
opciones_edge = Options()

# Configuramos el navegador para usar el motor de Chromium (que es el motor de Google Chrome).
opciones_edge.use_chromium = True

# Indicamos la ubicación del archivo del controlador de Microsoft Edge.
# Este archivo permite que el programa controle el navegador Edge.
servicio_edge = Service('C:\\Users\\William Monroy\\Desktop\\CODE\\msedgedriver.exe')

# Creamos un objeto que representa el navegador Microsoft Edge.
# Le decimos al programa dónde encontrar el controlador y qué opciones usar.
driver = webdriver.Edge(service=servicio_edge, options=opciones_edge)


# Creamos un nuevo libro de Excel en blanco.
libro_excel = openpyxl.Workbook()

# Verificamos si hay una hoja con el nombre 'Sheet' en el libro de Excel.
if 'Sheet' in libro_excel.sheetnames:
    # Si existe una hoja llamada 'Sheet', la eliminamos del libro de Excel.
    libro_excel.remove(libro_excel['Sheet'])
    


# La función `mostrar_carga` muestra una animación de carga en la consola.
# Se actualiza del 0% al 100% durante el tiempo especificado en segundos.
# La animación se muestra como un porcentaje que incrementa gradualmente.
# La función hace una pausa corta entre cada actualización para que el usuario
# pueda ver el progreso. Al final, muestra un mensaje indicando que la carga
# ha completado al 100%.
                        
def carga_completada(tiempo_total):
    
# Define el número total de pasos para la carga (100 pasos representan el 100% de la carga).
    numero_de_pasos = 100
    
# Calculamos el tiempo a esperar entre cada paso del indicador de carga.
    #tiempo_total es el tiempo total durante el cual debe mostrarse el indicador de carga.
    #numero_de_pasos es el número total de pasos (en este caso, 100 pasos).
    #tiempo_por_paso es el tiempo que debe esperar entre cada actualización del indicador.
    # Esto se obtiene dividiendo el tiempo total entre el número de pasos.
    tiempo_por_paso = tiempo_total / numero_de_pasos

# Iteramos sobre cada paso del indicador de carga.
    # Bucle para actualizar el indicador de carga desde 0% hasta 100%.
    for paso in range(numero_de_pasos + 1):
        # Calcula el porcentaje completado en el paso actual.
        porcentaje_actual = paso
        
# Imprimimos el porcentaje de carga en la misma línea de la consola.
        # El carácter \r mueve el cursor al principio de la línea para sobrescribir el texto anterior.
        sys.stdout.write(f"\rCARGA: {porcentaje_actual}%")
        
# Forzamos la actualización del texto en la consola.
        # Asegura que el texto se muestre de inmediato en la consola.
        sys.stdout.flush()

# Esperamos el tiempo calculado antes de actualizar al siguiente paso.
        # Pausa la ejecución del programa por el tiempo calculado antes de actualizar el próximo paso.
        # Esto crea el efecto de animación de carga.
        time.sleep(tiempo_por_paso)

# Una vez completada la carga, mostramos un mensaje final indicando que la carga está completa.
    # El carácter \r mueve el cursor al principio de la línea y el texto adicional borra cualquier residual de la línea anterior.
    sys.stdout.write("\rCARGA: 100% ")
    print("¡LA CARGA SE COMPLETÓ EXITOSAMENTE!\n")
    

# Forzamos la actualización del mensaje final en la consola.
    # Asegura que el mensaje final se muestre de inmediato en la consola.
    sys.stdout.flush()


#carga_completada(10)
    # Llama a la función con un tiempo total de 10 segundos
    # Esto hará que el indicador de carga se actualice cada 0.1 segundos (10 segundos / 100 pasos)
    # durante 10 segundos, mostrando el progreso desde 0% hasta 100%.


    
# Definimos una función de espera de 1 segundo para la selección del valor correspondiente ala ciudad
def pausa_ciudad_segundos():
    print("POR FAVOR, ESPERA MIENTRAS SE PROCESA TU SELECCIÓN DE CIUDAD...")
    time.sleep(1)
    

# Definimos una función de espera de 1 segundo para la selección del valor correspondiente al cine
def pausa_cine_segundos():
    print("POR FAVOR, ESPERA MIENTRAS SE PROCESA TU SELECCIÓN DE CINE...")
    time.sleep(1)



# Esta función intenta cerrar una ventana emergente que puede aparecer en el navegador.
# La función realiza los siguientes pasos:
# 1. Espera a que la ventana emergente sea visible en la página web.
# 2. Busca el botón de cerrar de la ventana emergente usando su identificador (ID).
# 3. Hace clic en el botón para cerrar la ventana emergente.
# 4. Imprime un mensaje indicando que se está cerrando la ventana emergente.
# Si ocurre algún error durante este proceso (como si el botón no se encuentra o no se puede hacer clic),
# se captura el error y se imprime un mensaje indicando el problema.
def cerrar_ventana_emergente(driver, espera):

    try:
        # Espera hasta que un elemento con la clase 'tk' esté presente en la página.
        espera.until(EC.presence_of_element_located((By.CSS_SELECTOR, "div.tk")))

        # Busca el botón para cerrar la ventana emergente usando su ID 'takeover-close'.
        boton_cerrar = driver.find_element(By.ID, "takeover-close")

        # Simula un clic en el botón para cerrar la ventana emergente.
        boton_cerrar.click()
        print("\nESTAMOS CERRANDO LA VENTANA EMERGENTE. POR FAVOR, ESPERE...")

        #Añade una pausa de cinco segungos mientras se cierra la ventana emergente
        time.sleep(5)
        print("¡LA VENTANA EMERGENTE SE HA CERRADO CON ÉXITO.!\n")

    except Exception as e:
        # Si ocurre un error al intentar cerrar la ventana, imprime un mensaje con los detalles del error.
        print(f"\nNO SE PUDO CERRAR LA VENTANA EMERGENTE: {e}\n")





# Esta función permite seleccionar una ciudad y un cine de una página web que usa un menú desplegable.
# Toma como entrada el navegador web (`driver`), un objeto para gestionar el tiempo de espera (`espera`),
# el valor de la ciudad que queremos seleccionar (`valor_ciudad`), y el valor del cine que queremos seleccionar (`valor_cine`).
# Luego, espera a que aparezcan los menús desplegables correspondientes, selecciona la ciudad y el cine según los valores dados,
# y devuelve el nombre del cine seleccionado en mayúsculas.
def elegir_ciudad_cine(driver, espera, valor_ciudad, valor_cine):

    # Pausa la ejecución del script para permitir que la página web se cargue completamente antes de interactuar.
    pausa_ciudad_segundos()
    carga_completada(2)

    # Espera hasta que el elemento HTML con el id "ciudad" esté presente en la página.
    # Este elemento es un menú desplegable (select) que permite al usuario seleccionar una ciudad.
    VARIABLE_HTML_select_ciudad = espera.until(EC.presence_of_element_located((By.ID, "ciudad")))

    # Crea un objeto Select para manejar el menú desplegable de ciudades.
    select_ciudad = Select(VARIABLE_HTML_select_ciudad)

    # Selecciona la opción de ciudad deseada en el menú desplegable usando el valor proporcionado (valor_ciudad).
    # En el HTML, esta opción se identifica por el atributo "value" del elemento <option> dentro del <select id="ciudad">.
    select_ciudad.select_by_value(valor_ciudad)

    # Espera hasta que todas las opciones en el menú desplegable de cines estén presentes en la página.
    # Esto se asegura esperando a que se carguen todas las opciones dentro del elemento <select id="cine">.
    espera.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, "select#cine option")))

    # Pausa la ejecución del script para dar tiempo a que las opciones de cine se carguen después de seleccionar la ciudad.
    pausa_cine_segundos()
    carga_completada(2)

    # Espera hasta que el elemento HTML con el id "cine" esté presente en la página.
    # Este elemento es otro menú desplegable (select) que permite al usuario seleccionar un cine específico.
    VARIABLE_HTML_select_cine = espera.until(EC.presence_of_element_located((By.ID, "cine")))

    # Crea un objeto Select para manejar el menú desplegable de cines.
    select_cine = Select(VARIABLE_HTML_select_cine)

    # Selecciona la opción de cine deseada en el menú desplegable usando el valor proporcionado (valor_cine).
    # En el HTML, esta opción se identifica por el atributo "value" del elemento <option> dentro del <select id="cine">.
    select_cine.select_by_value(valor_cine)

    # Obtiene la opción actualmente seleccionada en el menú desplegable de cines.
    # Esto accede al texto visible del elemento <option> que está seleccionado en el <select id="cine">.
    opcion_seleccionada = select_cine.first_selected_option

    # Obtiene el texto de la opción seleccionada (nombre del cine) y lo convierte a mayúsculas.
    nombre_cine = opcion_seleccionada.text.upper()

    # Devuelve el nombre del cine seleccionado.
    return nombre_cine




# Esta función copia datos sobre películas y formatos desde una fuente de datos a una hoja de cálculo de Excel.
# Toma como entrada el nombre del cine (`nombre_cine`) y un objeto que representa el archivo de Excel (`libro_excel`).
# La función realiza las siguientes tareas:
# 1. Crea una nueva hoja en el archivo de Excel con un nombre basado en el nombre del cine.
# 2. Limita el nombre de la hoja a un máximo de 31 caracteres y asegura que el nombre sea único dentro del archivo.
# 3. Copia datos de películas y formatos a las columnas correspondientes en la hoja de Excel:
#    - Columna E: Títulos de las películas.
#    - Columna F: Horas de proyección.
#    - Columna G: Idiomas o formatos (como 3D o 2D).
#    - Columna H: Formatos adicionales de la película.
# 4. Agrega la fecha actual en la columna A, el país en la columna B y el nombre del cine en la columna C.
# Esta función organiza y almacena la información de manera estructurada para su análisis y visualización en Excel.
def copiar_datos_a_excel(nombre_cine, libro_excel):

# Simplificar el nombre de la hoja para que no sea demasiado largo
    # Excel solo permite nombres de hojas que tengan hasta 31 caracteres.
    # Si el nombre del cine es más largo que eso, solo tomamos los primeros 31 caracteres.
    nombre_cine_corto = (nombre_cine[:31] if len(nombre_cine) > 31 else nombre_cine).upper()
    
# Asegurarse de que el nombre de la hoja sea único
    # Excel no permite tener dos hojas con el mismo nombre en un mismo archivo.
    # Si ya hay una hoja con el mismo nombre, modificamos el nombre para que sea único.
    while nombre_cine_corto in libro_excel.sheetnames:
        # Añadimos un número al final del nombre para hacerlo único.
        # Por ejemplo, si ya existe una hoja llamada "Cine1", la nueva hoja se llamará "Cine1_2" si hay más de una hoja.
        nombre_cine_corto = (nombre_cine_corto[:30] + '_' + str(len(libro_excel.sheetnames))).upper()
    
#Crear una nueva hoja en el archivo de Excel
    # Usamos el nombre ajustado para crear una nueva hoja en el archivo.
    hoja_excel = libro_excel.create_sheet(title=nombre_cine_corto)

    # Crea una lista vacía para almacenar los nombres de las películas que encontraremos en la página web.
    lista_peliculas = []

    # Crea un diccionario vacío para almacenar los formatos disponibles para cada hora de cada película.
    lista_formatos_por_hora = {}

    # Crea un diccionario vacío para almacenar los formatos disponibles para cada película.
    formato_por_pelicula = {}

    # Aquí comenzamos a buscar información en la página web.
    # Buscamos todos los elementos en la página que contienen información general sobre las películas.
    # En el código HTML, buscamos todos los elementos que tienen la clase 'informacion-general', que suele ser un contenedor con información de la película.
    for VARIABLE_HTML_div in driver.find_elements(By.CSS_SELECTOR, "div.informacion-general"):
        
        # Dentro de cada contenedor encontrado, buscamos el título de la película.
        # En el código HTML, el título de la película está dentro de una etiqueta <h3> dentro del div.
        VARIABLE_HTML_h3 = VARIABLE_HTML_div.find_element(By.TAG_NAME, "h3")
        pelicula = VARIABLE_HTML_h3.text  # Extraemos el nombre de la película.
        
        # Añadimos el nombre de la película a nuestra lista de películas.
        lista_peliculas.append(pelicula)
        
        # Buscamos todas las imágenes dentro del mismo contenedor.
        # En el código HTML, estas imágenes están dentro de la misma <div> que contiene la información de la película.
        img_tags = VARIABLE_HTML_div.find_elements(By.TAG_NAME, "img")
        
        # Creamos una lista de formatos a partir de los atributos 'alt' de las imágenes.
        # Solo tomamos los atributos 'alt' que no estén vacíos.
        formatos = [img.get_attribute("alt") for img in img_tags if img.get_attribute("alt")]
        
        # Guardamos la lista de formatos en nuestro diccionario usando el nombre de la película como clave.
        formato_por_pelicula[pelicula] = formatos

    # Ahora, configuramos la primera fila de la hoja de Excel para que tenga el título "TITULO".
    # En el archivo de Excel, la celda en la columna 'E' y fila 1 (E1) se usa para colocar el encabezado "TITULO".
    hoja_excel['E1'] = "TITULO"

    # Empezamos en la segunda fila y ponemos cada nombre de película en la columna 'E'.
    # En Excel, colocamos los nombres de las películas en la columna E, comenzando desde la fila 2 para evitar el encabezado.
    for idx, pelicula in enumerate(lista_peliculas, start=2):
        hoja_excel[f'E{idx}'] = pelicula

    # Creamos un nuevo diccionario vacío para almacenar las horas disponibles para cada película.
    # En el futuro, usaremos este diccionario para relacionar cada película con sus horarios disponibles.
    lista_horas_por_pelicula = {}


    try:
        # Recorremos cada sección de información sobre películas en la página web.
        # Cada sección se encuentra en un elemento HTML con la clase 'informacion-general'.
        for VARIABLE_HTML_div in driver.find_elements(By.CSS_SELECTOR, "div.informacion-general"):
            
            # Dentro de cada sección, encontramos el título de la película.
            # El título se encuentra dentro de una etiqueta HTML <h3>.
            VARIABLE_HTML_h3 = VARIABLE_HTML_div.find_element(By.TAG_NAME, "h3")
            pelicula = VARIABLE_HTML_h3.text  # Extraemos el nombre de la película.
            
            # Buscamos todas las partes que contienen información sobre formatos en la misma sección.
            # Estos formatos están dentro de elementos HTML con la clase 'formato'.
            formato_divs = VARIABLE_HTML_div.find_elements(By.CSS_SELECTOR, "div.formato")
            
            # Creamos una lista para almacenar las horas en las que la película está disponible.
            lista_horas = []

            # Recorremos cada parte que contiene información sobre un formato.
            for formato_div in formato_divs:
                # Extraemos el nombre del formato (por ejemplo, 2D, 3D) de una etiqueta <p> dentro de la parte del formato.
                formato_nombre = formato_div.find_element(By.CSS_SELECTOR, "p.formato-nombre").text
                
                # Buscamos todas las horas disponibles para ese formato.
                # Estas horas están en elementos HTML dentro de <div class='horas'> y son enlaces <a> con un <p> que contiene el texto de la hora.
                horas = formato_div.find_elements(By.CSS_SELECTOR, "div.horas a p")

                # Recorremos cada hora disponible.
                for hora in horas:
                    # Extraemos el texto de la hora (por ejemplo, 18:00, 20:30).
                    hora_text = hora.text
                    
                    # Obtenemos el enlace asociado a esa hora (por ejemplo, para más detalles o reserva).
                    # El enlace está en el elemento HTML padre del <p> que contiene la hora.
                    link = hora.find_element(By.XPATH, "..").get_attribute("href")
                    
                    # Añadimos la hora a nuestra lista de horas disponibles para la película.
                    lista_horas.append(hora_text)
                    
                    # Guardamos el nombre del formato para cada hora en un diccionario.
                    lista_formatos_por_hora[hora_text] = formato_nombre

            # Guardamos la lista de horas disponibles para cada película en un diccionario.
            # La clave es el nombre de la película y el valor es la lista de horas.
            lista_horas_por_pelicula[pelicula] = lista_horas
    except Exception as e:
        # Si ocurre un error al extraer la información, mostramos un mensaje de error.
        print(f"Error al extraer horas y formatos de las películas: {e}")


    # Asignamos títulos a las columnas de la hoja de Excel para que cada columna tenga una etiqueta clara.
    hoja_excel['F1'] = "HORA"      # La columna F tendrá el título "HORA", que indica el horario de las proyecciones.
    hoja_excel['G1'] = "IDIOMA"    # La columna G tendrá el título "IDIOMA", que indica el idioma en el que se proyecta la película.
    hoja_excel['H1'] = "FORMATO"   # La columna H tendrá el título "FORMATO", que muestra el tipo de formato (como 2D o 3D).
    hoja_excel['D1'] = "NOMBRE CINE" # La columna D tendrá el título "NOMBRE CINE", que indica el nombre del cine.

    # Empezamos a escribir datos en la fila 2 de la hoja de Excel (la fila 1 ya contiene los títulos de las columnas).
    fila = 2

    # Recorremos cada película y sus horarios de proyección.
    for pelicula, lista_horas in lista_horas_por_pelicula.items():
        # Para cada hora en la lista de horas disponibles para la película,
        for hora in lista_horas:
            # Escribimos el nombre del cine en la columna D para la fila actual.
            hoja_excel[f'D{fila}'] = nombre_cine
            # Escribimos el nombre de la película en la columna E para la fila actual.
            hoja_excel[f'E{fila}'] = pelicula
            # Escribimos la hora de la proyección en la columna F para la fila actual.
            hoja_excel[f'F{fila}'] = hora

            # Buscamos el formato de idioma asociado con esta hora en el diccionario.
            formato_idioma = lista_formatos_por_hora.get(hora, "")
            # Escribimos el formato de idioma en la columna G para la fila actual.
            hoja_excel[f'G{fila}'] = formato_idioma
            
            # Verificamos si el formato de idioma contiene "3D" o "DIG".
            # Si el formato incluye "3D", ponemos "3D" en la columna H para esa fila.
            # Si el formato incluye "DIG", ponemos "DIG" en la columna H para esa fila.
            if "3D" in formato_idioma or "DIG" in formato_idioma:
                hoja_excel[f'H{fila}'] = "3D" if "3D" in formato_idioma else "DIG"
                # Eliminamos "3D" o "DIG" del texto del formato de idioma para que quede solo el nombre del idioma.
                formato_idioma = formato_idioma.replace("3D", "").replace("DIG", "").strip()
                # Escribimos el nombre del idioma (sin "3D" o "DIG") en la columna G.
                hoja_excel[f'G{fila}'] = formato_idioma

            
            # Si la columna H (FORMATO) está vacía, la llenamos con "2D" por defecto.
            if not hoja_excel[f'H{fila}'].value:
                hoja_excel[f'H{fila}'] = "2D"
            
            # Obtenemos una lista de formatos adicionales para esta película.
            formatos_adicionales = formato_por_pelicula.get(pelicula, [])
            
            # Si hay formatos adicionales, los combinamos con el formato existente en la columna H.
            if formatos_adicionales:
                # Obtenemos el formato que ya está en la columna H para la fila actual. Si está vacío, usamos una cadena vacía.
                formato_existente = hoja_excel[f'H{fila}'].value or ""
                
                # Combinamos los formatos adicionales con el formato existente. Si hay varios formatos, los unimos en una sola cadena separada por espacios.
                hoja_excel[f'H{fila}'] = " ".join([f"{formato}" for formato in formatos_adicionales] + [formato_existente])

            # Pasamos a la siguiente fila para continuar escribiendo los datos de la próxima proyección.
            fila += 1


    # Establece el título de la primera columna en la hoja de Excel como 'FECHA'.
    hoja_excel['A1'] = 'FECHA'
    
    # Obtiene el número total de filas en la hoja de Excel que contienen datos.
    max_filas = hoja_excel.max_row
    
    # Obtiene la fecha actual en formato 'YYYY-MM-DD' (por ejemplo, '2024-08-12').
    fecha_actual = datetime.now().strftime('%Y-%m-%d')

    # Recorre todas las filas de la hoja de Excel, comenzando desde la fila 2 hasta la última fila con datos.
    for fila in range(2, max_filas + 1):
        # Establece el valor de la celda en la columna A (FECHA) para la fila actual con la fecha actual.
        hoja_excel[f'A{fila}'] = fecha_actual





    # Creamos un diccionario llamado 'dominio_a_pais'
    # Un diccionario es una estructura de datos que guarda pares de clave-valor.
    # En este caso, la clave es una extensión de dominio (como 'sv') y el valor es el nombre del país correspondiente (como 'EL SALVADOR').
    dominio_a_pais = {
        'sv': 'EL SALVADOR',      # La clave 'sv' se asocia con el valor 'EL SALVADOR'.
        'gt': 'GUATEMALA',        # La clave 'gt' se asocia con el valor 'GUATEMALA'.
        'cr': 'COSTA RICA',       # La clave 'cr' se asocia con el valor 'COSTA RICA'.
        'pa': 'PANAMÁ'            # La clave 'pa' se asocia con el valor 'PANAMÁ'.
    }


    # Esta función determina el país al que pertenece una URL basada en el dominio de la página web.
    # Toma como entrada una cadena de texto (`url`) que representa la dirección web.
    # La función realiza los siguientes pasos:
    # 1. Extrae la parte del dominio de la URL después de "https://" y antes del primer "/".
    # 2. Identifica la extensión del dominio, que es la parte final del dominio después del último punto (por ejemplo, "sv" en "cinepolis.com.sv").
    # 3. Usa un diccionario (`dominio_a_pais`) para mapear esta extensión de dominio a un nombre de país específico.
    # 4. Devuelve el nombre del país asociado a la extensión del dominio. Si la extensión no está en el diccionario, devuelve "DESCONOCIDO".
    # Esta función es útil para clasificar o identificar la ubicación geográfica basada en el dominio de la página web.
    def obtener_pais(url):

        # Extrae el dominio de la URL, eliminando el protocolo (http://, https://) y cualquier ruta adicional.
        dominio = url.split('//')[-1].split('/')[0]
        # Obtiene la extensión del dominio, por ejemplo, "sv" de "cinepolis.com.sv".
        extension = dominio.split('.')[-1]
        # Busca la extensión en el diccionario para obtener el país. Si no se encuentra, devuelve 'DESCONOCIDO'.
        return dominio_a_pais.get(extension, 'DESCONOCIDO')

    # Obtiene el nombre del país basándose en la URL.
    nombrePais = obtener_pais(url)

    # Establece el encabezado "PAIS" en la celda B1 de la hoja de Excel.
    hoja_excel["B1"] = "PAIS"
    # Recorre todas las filas con datos en la hoja de Excel, comenzando desde la fila 2.
    for fila in range(2, max_filas + 1):
        # Establece el valor de la celda en la columna B (PAIS) para la fila actual con el nombre del país.
        hoja_excel[f"B{fila}"] = nombrePais



    
    # Extrae el nombre del cine de la URL y lo convierte a mayúsculas.
    # Supone que la URL tiene el formato "https://nombre-del-cine.com.gt/..." 
    cadena_de_cine = url.split("https://")[1].split(".com.gt/")[0].upper()

    # Establece el encabezado "CINE" en la celda C1 de la hoja de Excel.
    hoja_excel["C1"] = "CINE"

    # Recorre todas las filas con datos en la hoja de Excel, comenzando desde la fila 2.
    for fila in range(2, max_filas + 1):
        # Establece el valor de la celda en la columna C (CINE) para la fila actual con el nombre del cine extraído.
        hoja_excel[f"C{fila}"] = cadena_de_cine





# Esta función elimina iconos o símbolos no deseados de la columna H en el archivo Excel.
# La columna H se selecciona dentro del archivo de Excel que se pasa como argumento (`libro_excel`).
# La función realiza lo siguiente:
# 1. Accede a la hoja activa del libro de Excel.
# 2. Recorre cada celda en la columna H.
# 3. Reemplaza cualquier icono o símbolo específico con una cadena vacía, dejando solo el texto limpio en la celda.
# Este proceso es útil para limpiar los datos antes de analizarlos o guardarlos, asegurando que solo se mantenga el texto relevante.
def eliminar_icono_de_columna_h(libro_excel):

    # Recorremos cada hoja en el archivo de Excel.
    # Cada hoja en un libro de Excel es una página donde se pueden encontrar varias tablas o datos.
    for hoja in libro_excel.worksheets:
        # Iteramos sobre cada fila en la columna H (la columna número 8), comenzando desde la segunda fila.
        # Usamos `min_col=8` y `max_col=8` para centrarnos solo en la columna H.
        # `min_row=2` significa que comenzamos en la segunda fila, generalmente porque la primera fila contiene encabezados.
        for fila in hoja.iter_rows(min_col=8, max_col=8, min_row=2):
            # `fila` es una lista que contiene las celdas de la columna H para una fila específica.
            # Tomamos la primera celda en esta lista, que corresponde a la columna H.
            celda = fila[0]
            # Verificamos si la celda tiene un valor y si ese valor es una cadena de texto.
            # `celda.value` contiene el contenido de la celda, y `isinstance(celda.value, str)` asegura que es texto.
            if celda.value and isinstance(celda.value, str):
                # `replace("icono-", "")` busca la palabra "icono-" en el texto y la reemplaza con una cadena vacía.
                # Esto elimina "icono-" del contenido de la celda si está presente.
                celda.value = celda.value.replace("icono-", "")



# Esta función elimina palabras repetidas en las celdas de la columna H en un archivo Excel.
# La columna H se selecciona dentro del archivo de Excel que se pasa como argumento (`libro_excel`).
# La función realiza los siguientes pasos:
# 1. Accede a la hoja activa del libro de Excel.
# 2. Recorre cada celda en la columna H.
# 3. En cada celda, identifica y elimina las palabras que se repiten.
# 4. Guarda el contenido limpio, con palabras únicas, en la misma celda.
# Este proceso ayuda a limpiar los datos en la columna H para que no haya duplicados y sea más fácil analizarlos.
def eliminar_palabras_repetidas_columna_h(libro_excel):

    # Recorremos cada hoja en el archivo de Excel.
    # Esto es necesario porque queremos aplicar la limpieza a todas las hojas en el libro.
    for hoja in libro_excel.worksheets:
        # Iteramos sobre cada fila en la columna H (la columna número 8), comenzando desde la segunda fila.
        # Usamos `min_col=8` y `max_col=8` para centrarnos solo en la columna H.
        # `min_row=2` se usa porque la primera fila generalmente contiene encabezados.
        for fila in hoja.iter_rows(min_col=8, max_col=8, min_row=2):
            # `fila` es una lista que contiene las celdas de la columna H para una fila específica.
            # Tomamos la primera celda en esta lista, que corresponde a la columna H.
            celda = fila[0]
            # Verificamos si la celda tiene un valor y si ese valor es una cadena de texto.
            # `celda.value` contiene el contenido de la celda, y `isinstance(celda.value, str)` asegura que sea texto.
            if celda.value and isinstance(celda.value, str):
                # Dividimos el texto en palabras usando el espacio como separador.
                palabras = celda.value.split()
                # Usamos `dict.fromkeys(palabras)` para eliminar palabras duplicadas.
                # Convertimos la lista de palabras en un diccionario (donde las claves son únicas) y luego de vuelta a una lista.
                palabras_unicas = list(dict.fromkeys(palabras))
                # Volvemos a unir las palabras únicas en una cadena de texto, separadas por espacios.
                celda.value = ' '.join(palabras_unicas)



# Esta función convierte todo el texto en la columna H de un archivo Excel a letras mayúsculas.
# La columna H se selecciona dentro del archivo de Excel que se pasa como argumento (`libro_excel`).
# La función realiza los siguientes pasos:
# 1. Accede a la hoja activa del libro de Excel.
# 2. Recorre cada celda en la columna H.
# 3. Toma el texto de cada celda y lo convierte a letras mayúsculas.
# 4. Guarda el texto convertido en la misma celda.
# Esto asegura que todo el texto en la columna H esté en mayúsculas, lo que puede ayudar a mantener la consistencia y facilitar la lectura.
def convertir_a_mayusculas_columna_h(libro_excel):

    # Recorre todas las hojas en el libro de Excel.
    # 'libro_excel.worksheets' devuelve una lista de todas las hojas en el libro.
    for hoja in libro_excel.worksheets:
        # Recorre cada fila en la columna H (columna 8), comenzando desde la segunda fila para evitar el encabezado.
        # 'iter_rows' permite iterar sobre un rango específico de celdas. 'min_col=8' y 'max_col=8' indican que solo se debe considerar la columna H.
        # 'min_row=2' asegura que se empiece desde la segunda fila para omitir el encabezado.
        for fila in hoja.iter_rows(min_col=8, max_col=8, min_row=2):
            # Obtiene la celda en la columna H para la fila actual.
            # 'fila[0]' accede a la primera celda de la fila, ya que solo estamos considerando una columna (H).
            celda = fila[0]
            
            # Verifica si la celda tiene un valor y si el valor es una cadena de texto.
            # 'celda.value' obtiene el valor de la celda y 'isinstance(celda.value, str)' asegura que el valor sea texto.
            if celda.value and isinstance(celda.value, str):
                # Cambia el texto de la celda a mayúsculas usando el método 'upper()'.
                # 'celda.value.upper()' convierte todo el texto de la celda a letras grandes.
                celda.value = celda.value.upper()



# Esta función convierte todo el texto en la columna E de un archivo Excel a letras mayúsculas.
# La columna E se selecciona dentro del archivo de Excel que se pasa como argumento (`libro_excel`).
# La función realiza los siguientes pasos:
# 1. Accede a la hoja activa del libro de Excel.
# 2. Recorre cada celda en la columna E.
# 3. Toma el texto de cada celda y lo convierte a letras mayúsculas.
# 4. Guarda el texto convertido en la misma celda.
# Esto asegura que todo el texto en la columna E esté en mayúsculas, lo que ayuda a mantener la uniformidad en los datos y facilita su lectura.
def convertir_a_mayusculas_columna_e(libro_excel):

    # Recorre todas las hojas en el libro de Excel.
    # 'libro_excel.worksheets' devuelve una lista de todas las hojas en el libro.
    for hoja in libro_excel.worksheets:
        # Recorre cada fila en la columna E (columna 5), comenzando desde la segunda fila para evitar el encabezado.
        # 'iter_rows' permite iterar sobre un rango específico de celdas. 'min_col=5' y 'max_col=5' indican que solo se debe considerar la columna E.
        # 'min_row=2' asegura que se empiece desde la segunda fila para omitir el encabezado.
        for fila in hoja.iter_rows(min_col=5, max_col=5, min_row=2):
            # Obtiene la celda en la columna E para la fila actual.
            # 'fila[0]' accede a la primera celda de la fila, ya que solo estamos considerando una columna (E).
            celda = fila[0]
            
            # Verifica si la celda tiene un valor y si el valor es una cadena de texto.
            # 'celda.value' obtiene el valor de la celda y 'isinstance(celda.value, str)' asegura que el valor sea texto.
            if celda.value and isinstance(celda.value, str):
                # Cambia el texto de la celda a mayúsculas usando el método 'upper()'.
                # 'celda.value.upper()' convierte todo el texto de la celda a letras grandes.
                celda.value = celda.value.upper()


try:
    # Define la URL del sitio web al que se desea acceder.
    url = "https://cinepolis.com.gt/"
    
    # Abre la página web especificada en la URL usando el navegador controlado por Selenium.
    driver.get(url)
    
    # Crea una instancia de WebDriverWait con un tiempo de espera máximo de 10 segundos.
    espera = WebDriverWait(driver, 10)
    
    # Llama a la función `cerrar_ventana_emergente` para manejar y cerrar cualquier ventana emergente.
    cerrar_ventana_emergente(driver, espera)
    
#1  #Cinépolis Cayalá

    # Llamada a la función elegir_ciudad_cine para seleccionar una ciudad y un cine específicos
    # valor_ciudad = "42" corresponde al atributo 'value' de una opción en el menú desplegable de ciudades
    # en el HTML, se verá algo como esto:
    # <select id="ciudad">
    #   <option value="42">Ciudad Ejemplo</option>
    #   <!-- Otras opciones -->
    # </select>
    #
    # valor_cine = "417" corresponde al atributo 'value' de una opción en el menú desplegable de cines
    # en el HTML, se verá algo como esto:
    # <select id="cine">
    #   <option value="417">Cine Ejemplo</option>
    #   <!-- Otras opciones -->
    # </select>

    nombre_cine = elegir_ciudad_cine(driver, espera, "42", "417")

   # Se imprime el nombre del cine seleccionado para confirmación.
    print(f"\n\n\n\n\n\nNOMBRE DEL CINE SELECCIONADO (1/17): {nombre_cine}.")

    # Informamos al usuario que el proceso de copia de datos desde la página web está en curso.
    print("ESTAMOS COPIANDO LOS DATOS DE LA PÁGINA WEB. POR FAVOR, ESPERE...")
    # Se realiza una pausa de diez segundos para asegurar que el proceso se complete correctamente y minimizar el riesgo de errores.
    carga_completada(10)

    # Se llama a la función `copiar_datos_a_excel`, pasándole el nombre del cine y el libro de Excel para que los datos sean copiados a la hoja de cálculo.
    copiar_datos_a_excel(nombre_cine, libro_excel)





#2  #Cinépolis El Frutal

    # Crea una instancia de WebDriverWait con un tiempo de espera máximo de 10 segundos.
    espera = WebDriverWait(driver, 10)

    # Llamada a la función elegir_ciudad_cine para seleccionar una ciudad y un cine específicos
    # valor_ciudad = "42" corresponde al atributo 'value' de una opción en el menú desplegable de ciudades
    # en el HTML, se verá algo como esto:
    # <select id="ciudad">
    #   <option value="42">Ciudad Ejemplo</option>
    #   <!-- Otras opciones -->
    # </select>
    #
    # valor_cine = "511" corresponde al atributo 'value' de una opción en el menú desplegable de cines
    # en el HTML, se verá algo como esto:
    # <select id="cine">
    #   <option value="511">Cine Ejemplo</option>
    #   <!-- Otras opciones -->
    # </select>

    nombre_cine = elegir_ciudad_cine(driver, espera, "42", "511")


   # Se imprime el nombre del cine seleccionado para confirmación.
    print(f"\n\n\n\n\n\nNOMBRE DEL CINE SELECCIONADO (2/17): {nombre_cine}.")

    # Informamos al usuario que el proceso de copia de datos desde la página web está en curso.
    print("ESTAMOS COPIANDO LOS DATOS DE LA PÁGINA WEB. POR FAVOR, ESPERE...")
    # Se realiza una pausa de diez segundos para asegurar que el proceso se complete correctamente y minimizar el riesgo de errores.
    carga_completada(10)

    # Se llama a la función `copiar_datos_a_excel`, pasándole el nombre del cine y el libro de Excel para que los datos sean copiados a la hoja de cálculo.
    copiar_datos_a_excel(nombre_cine, libro_excel)



#3  #Cinépolis Galerías Miraflores

    # Crea una instancia de WebDriverWait con un tiempo de espera máximo de 10 segundos.
    espera = WebDriverWait(driver, 10)

    # Llamada a la función elegir_ciudad_cine para seleccionar una ciudad y un cine específicos
    # valor_ciudad = "42" corresponde al atributo 'value' de una opción en el menú desplegable de ciudades
    # en el HTML, se verá algo como esto:
    # <select id="ciudad">
    #   <option value="42">Ciudad Ejemplo</option>
    #   <!-- Otras opciones -->
    # </select>
    #
    # valor_cine = "315" corresponde al atributo 'value' de una opción en el menú desplegable de cines
    # en el HTML, se verá algo como esto:
    # <select id="cine">
    #   <option value="315">Cine Ejemplo</option>
    #   <!-- Otras opciones -->
    # </select>

    nombre_cine = elegir_ciudad_cine(driver, espera, "42", "315")

   # Se imprime el nombre del cine seleccionado para confirmación.
    print(f"\n\n\n\n\n\nNOMBRE DEL CINE SELECCIONADO (3/17): {nombre_cine}.")

    # Informamos al usuario que el proceso de copia de datos desde la página web está en curso.
    print("ESTAMOS COPIANDO LOS DATOS DE LA PÁGINA WEB. POR FAVOR, ESPERE...")
    # Se realiza una pausa de diez segundos para asegurar que el proceso se complete correctamente y minimizar el riesgo de errores.
    carga_completada(10)

    # Se llama a la función `copiar_datos_a_excel`, pasándole el nombre del cine y el libro de Excel para que los datos sean copiados a la hoja de cálculo.
    copiar_datos_a_excel(nombre_cine, libro_excel)




#4  #Cinépolis Naranjo Mall

    # Crea una instancia de WebDriverWait con un tiempo de espera máximo de 10 segundos.
    espera = WebDriverWait(driver, 10)

    # Llamada a la función elegir_ciudad_cine para seleccionar una ciudad y un cine específicos
    # valor_ciudad = "42" corresponde al atributo 'value' de una opción en el menú desplegable de ciudades
    # en el HTML, se verá algo como esto:
    # <select id="ciudad">
    #   <option value="42">Ciudad Ejemplo</option>
    #   <!-- Otras opciones -->
    # </select>
    #
    # valor_cine = "493" corresponde al atributo 'value' de una opción en el menú desplegable de cines
    # en el HTML, se verá algo como esto:
    # <select id="cine">
    #   <option value="493">Cine Ejemplo</option>
    #   <!-- Otras opciones -->
    # </select>

    nombre_cine = elegir_ciudad_cine(driver, espera, "42", "493")


    # Se imprime el nombre del cine seleccionado para confirmación.
    print(f"\n\n\n\n\n\nNOMBRE DEL CINE SELECCIONADO (4/17): {nombre_cine}.")

    # Informamos al usuario que el proceso de copia de datos desde la página web está en curso.
    print("ESTAMOS COPIANDO LOS DATOS DE LA PÁGINA WEB. POR FAVOR, ESPERE...")
    # Se realiza una pausa de diez segundos para asegurar que el proceso se complete correctamente y minimizar el riesgo de errores.
    carga_completada(10)

    # Se llama a la función `copiar_datos_a_excel`, pasándole el nombre del cine y el libro de Excel para que los datos sean copiados a la hoja de cálculo.
    copiar_datos_a_excel(nombre_cine, libro_excel)

    







#5  #Cinépolis Oakland Mall

    # Crea una instancia de WebDriverWait con un tiempo de espera máximo de 10 segundos.
    espera = WebDriverWait(driver, 10)

    # Llamada a la función elegir_ciudad_cine para seleccionar una ciudad y un cine específicos
    # valor_ciudad = "42" corresponde al atributo 'value' de una opción en el menú desplegable de ciudades
    # en el HTML, se verá algo como esto:
    # <select id="ciudad">
    #   <option value="42">Ciudad Ejemplo</option>
    #   <!-- Otras opciones -->
    # </select>
    #
    # valor_cine = "145" corresponde al atributo 'value' de una opción en el menú desplegable de cines
    # en el HTML, se verá algo como esto:
    # <select id="cine">
    #   <option value="145">Cine Ejemplo</option>
    #   <!-- Otras opciones -->
    # </select>

    nombre_cine = elegir_ciudad_cine(driver, espera, "42", "145")


    # Se imprime el nombre del cine seleccionado para confirmación.
    print(f"\n\n\n\n\n\nNOMBRE DEL CINE SELECCIONADO (5/17): {nombre_cine}.")

    # Informamos al usuario que el proceso de copia de datos desde la página web está en curso.
    print("ESTAMOS COPIANDO LOS DATOS DE LA PÁGINA WEB. POR FAVOR, ESPERE...")
    # Se realiza una pausa de diez segundos para asegurar que el proceso se complete correctamente y minimizar el riesgo de errores.
    carga_completada(10)

    # Se llama a la función `copiar_datos_a_excel`, pasándole el nombre del cine y el libro de Excel para que los datos sean copiados a la hoja de cálculo.
    copiar_datos_a_excel(nombre_cine, libro_excel)

    





#6  #Cinépolis Parque Las Américas

    # Crea una instancia de WebDriverWait con un tiempo de espera máximo de 10 segundos.
    espera = WebDriverWait(driver, 10)

    # Llamada a la función elegir_ciudad_cine para seleccionar una ciudad y un cine específicos
    # valor_ciudad = "42" corresponde al atributo 'value' de una opción en el menú desplegable de ciudades
    # en el HTML, se verá algo como esto:
    # <select id="ciudad">
    #   <option value="42">Ciudad Ejemplo</option>
    #   <!-- Otras opciones -->
    # </select>
    #
    # valor_cine = "541" corresponde al atributo 'value' de una opción en el menú desplegable de cines
    # en el HTML, se verá algo como esto:
    # <select id="cine">
    #   <option value="541">Cine Ejemplo</option>
    #   <!-- Otras opciones -->
    # </select>

    nombre_cine = elegir_ciudad_cine(driver, espera, "42", "541")


    # Se imprime el nombre del cine seleccionado para confirmación.
    print(f"\n\n\n\n\n\nNOMBRE DEL CINE SELECCIONADO (6/17): {nombre_cine}.")

    # Informamos al usuario que el proceso de copia de datos desde la página web está en curso.
    print("ESTAMOS COPIANDO LOS DATOS DE LA PÁGINA WEB. POR FAVOR, ESPERE...")
    # Se realiza una pausa de diez segundos para asegurar que el proceso se complete correctamente y minimizar el riesgo de errores.
    carga_completada(10)

    # Se llama a la función `copiar_datos_a_excel`, pasándole el nombre del cine y el libro de Excel para que los datos sean copiados a la hoja de cálculo.
    copiar_datos_a_excel(nombre_cine, libro_excel)

    








#7  #Cinépolis Portales

    # Crea una instancia de WebDriverWait con un tiempo de espera máximo de 10 segundos.
    espera = WebDriverWait(driver, 10)

    # Llamada a la función elegir_ciudad_cine para seleccionar una ciudad y un cine específicos
    # valor_ciudad = "42" corresponde al atributo 'value' de una opción en el menú desplegable de ciudades
    # en el HTML, se verá algo como esto:
    # <select id="ciudad">
    #   <option value="42">Ciudad Ejemplo</option>
    #   <!-- Otras opciones -->
    # </select>
    #
    # valor_cine = "212" corresponde al atributo 'value' de una opción en el menú desplegable de cines
    # en el HTML, se verá algo como esto:
    # <select id="cine">
    #   <option value="212">Cine Ejemplo</option>
    #   <!-- Otras opciones -->
    # </select>

    nombre_cine = elegir_ciudad_cine(driver, espera, "42", "212")


    # Se imprime el nombre del cine seleccionado para confirmación.
    print(f"\n\n\n\n\n\nNOMBRE DEL CINE SELECCIONADO (7/17): {nombre_cine}.")

    # Informamos al usuario que el proceso de copia de datos desde la página web está en curso.
    print("ESTAMOS COPIANDO LOS DATOS DE LA PÁGINA WEB. POR FAVOR, ESPERE...")
    # Se realiza una pausa de diez segundos para asegurar que el proceso se complete correctamente y minimizar el riesgo de errores.
    carga_completada(10)

    # Se llama a la función `copiar_datos_a_excel`, pasándole el nombre del cine y el libro de Excel para que los datos sean copiados a la hoja de cálculo.
    copiar_datos_a_excel(nombre_cine, libro_excel)

    





#8  #Cinépolis Rus Mall

    # Crea una instancia de WebDriverWait con un tiempo de espera máximo de 10 segundos.
    espera = WebDriverWait(driver, 10)

    # Llamada a la función elegir_ciudad_cine para seleccionar una ciudad y un cine específicos
    # valor_ciudad = "42" corresponde al atributo 'value' de una opción en el menú desplegable de ciudades
    # en el HTML, se verá algo como esto:
    # <select id="ciudad">
    #   <option value="42">Ciudad Ejemplo</option>
    #   <!-- Otras opciones -->
    # </select>
    #
    # valor_cine = "586" corresponde al atributo 'value' de una opción en el menú desplegable de cines
    # en el HTML, se verá algo como esto:
    # <select id="cine">
    #   <option value="586">Cine Ejemplo</option>
    #   <!-- Otras opciones -->
    # </select>

    nombre_cine = elegir_ciudad_cine(driver, espera, "42", "586")


    # Se imprime el nombre del cine seleccionado para confirmación.
    print(f"\n\n\n\n\n\nNOMBRE DEL CINE SELECCIONADO (8/17): {nombre_cine}.")

    # Informamos al usuario que el proceso de copia de datos desde la página web está en curso.
    print("ESTAMOS COPIANDO LOS DATOS DE LA PÁGINA WEB. POR FAVOR, ESPERE...")
    # Se realiza una pausa de diez segundos para asegurar que el proceso se complete correctamente y minimizar el riesgo de errores.
    carga_completada(10)

    # Se llama a la función `copiar_datos_a_excel`, pasándole el nombre del cine y el libro de Excel para que los datos sean copiados a la hoja de cálculo.
    copiar_datos_a_excel(nombre_cine, libro_excel)

    





#9  #Cinépolis Sankris Mall

    # Crea una instancia de WebDriverWait con un tiempo de espera máximo de 10 segundos.
    espera = WebDriverWait(driver, 10)

    # Llamada a la función elegir_ciudad_cine para seleccionar una ciudad y un cine específicos
    # valor_ciudad = "42" corresponde al atributo 'value' de una opción en el menú desplegable de ciudades
    # en el HTML, se verá algo como esto:
    # <select id="ciudad">
    #   <option value="42">Ciudad Ejemplo</option>
    #   <!-- Otras opciones -->
    # </select>
    #
    # valor_cine = "559" corresponde al atributo 'value' de una opción en el menú desplegable de cines
    # en el HTML, se verá algo como esto:
    # <select id="cine">
    #   <option value="559">Cine Ejemplo</option>
    #   <!-- Otras opciones -->
    # </select>

    nombre_cine = elegir_ciudad_cine(driver, espera, "42", "559")


    # Se imprime el nombre del cine seleccionado para confirmación.
    print(f"\n\n\n\n\n\nNOMBRE DEL CINE SELECCIONADO (9/17): {nombre_cine}.")

    # Informamos al usuario que el proceso de copia de datos desde la página web está en curso.
    print("ESTAMOS COPIANDO LOS DATOS DE LA PÁGINA WEB. POR FAVOR, ESPERE...")
    # Se realiza una pausa de diez segundos para asegurar que el proceso se complete correctamente y minimizar el riesgo de errores.
    carga_completada(10)

    # Se llama a la función `copiar_datos_a_excel`, pasándole el nombre del cine y el libro de Excel para que los datos sean copiados a la hoja de cálculo.
    copiar_datos_a_excel(nombre_cine, libro_excel)

    











#10 #Cinépolis Santa Clara

    # Crea una instancia de WebDriverWait con un tiempo de espera máximo de 10 segundos.
    espera = WebDriverWait(driver, 10)

    # Llamada a la función elegir_ciudad_cine para seleccionar una ciudad y un cine específicos
    # valor_ciudad = "42" corresponde al atributo 'value' de una opción en el menú desplegable de ciudades
    # en el HTML, se verá algo como esto:
    # <select id="ciudad">
    #   <option value="42">Ciudad Ejemplo</option>
    #   <!-- Otras opciones -->
    # </select>
    #
    # valor_cine = "476" corresponde al atributo 'value' de una opción en el menú desplegable de cines
    # en el HTML, se verá algo como esto:
    # <select id="cine">
    #   <option value="476">Cine Ejemplo</option>
    #   <!-- Otras opciones -->
    # </select>

    nombre_cine = elegir_ciudad_cine(driver, espera, "42", "476")


    # Se imprime el nombre del cine seleccionado para confirmación.
    print(f"\n\n\n\n\n\nNOMBRE DEL CINE SELECCIONADO (10/17): {nombre_cine}.")

    # Informamos al usuario que el proceso de copia de datos desde la página web está en curso.
    print("ESTAMOS COPIANDO LOS DATOS DE LA PÁGINA WEB. POR FAVOR, ESPERE...")
    # Se realiza una pausa de diez segundos para asegurar que el proceso se complete correctamente y minimizar el riesgo de errores.
    carga_completada(10)

    # Se llama a la función `copiar_datos_a_excel`, pasándole el nombre del cine y el libro de Excel para que los datos sean copiados a la hoja de cálculo.
    copiar_datos_a_excel(nombre_cine, libro_excel)

    






#11 #Cinépolis VIP Oakland Mall

    # Crea una instancia de WebDriverWait con un tiempo de espera máximo de 10 segundos.
    espera = WebDriverWait(driver, 10)

    # Llamada a la función elegir_ciudad_cine para seleccionar una ciudad y un cine específicos
    # valor_ciudad = "42" corresponde al atributo 'value' de una opción en el menú desplegable de ciudades
    # en el HTML, se verá algo como esto:
    # <select id="ciudad">
    #   <option value="42">Ciudad Ejemplo</option>
    #   <!-- Otras opciones -->
    # </select>
    #
    # valor_cine = "146" corresponde al atributo 'value' de una opción en el menú desplegable de cines
    # en el HTML, se verá algo como esto:
    # <select id="cine">
    #   <option value="146">Cine Ejemplo</option>
    #   <!-- Otras opciones -->
    # </select>

    nombre_cine = elegir_ciudad_cine(driver, espera, "42", "146")

    
    # Se imprime el nombre del cine seleccionado para confirmación.
    print(f"\n\n\n\n\n\nNOMBRE DEL CINE SELECCIONADO (11/17): {nombre_cine}.")

    # Informamos al usuario que el proceso de copia de datos desde la página web está en curso.
    print("ESTAMOS COPIANDO LOS DATOS DE LA PÁGINA WEB. POR FAVOR, ESPERE...")
    # Se realiza una pausa de diez segundos para asegurar que el proceso se complete correctamente y minimizar el riesgo de errores.
    carga_completada(10)

    # Se llama a la función `copiar_datos_a_excel`, pasándole el nombre del cine y el libro de Excel para que los datos sean copiados a la hoja de cálculo.
    copiar_datos_a_excel(nombre_cine, libro_excel)

    





#12 #Cinépolis MC Jutiapa

    # Crea una instancia de WebDriverWait con un tiempo de espera máximo de 10 segundos.
    espera = WebDriverWait(driver, 10)

    # Llamada a la función elegir_ciudad_cine para seleccionar una ciudad y un cine específicos
    # valor_ciudad = "213" corresponde al atributo 'value' de una opción en el menú desplegable de ciudades
    # en el HTML, se verá algo como esto:
    # <select id="ciudad">
    #   <option value="213">Ciudad Ejemplo</option>
    #   <!-- Otras opciones -->
    # </select>
    #
    # valor_cine = "666" corresponde al atributo 'value' de una opción en el menú desplegable de cines
    # en el HTML, se verá algo como esto:
    # <select id="cine">
    #   <option value="666">Cine Ejemplo</option>
    #   <!-- Otras opciones -->
    # </select>

    nombre_cine = elegir_ciudad_cine(driver, espera, "213", "666")


    # Se imprime el nombre del cine seleccionado para confirmación.
    print(f"\n\n\n\n\n\nNOMBRE DEL CINE SELECCIONADO (12/17): {nombre_cine}.")

    # Informamos al usuario que el proceso de copia de datos desde la página web está en curso.
    print("ESTAMOS COPIANDO LOS DATOS DE LA PÁGINA WEB. POR FAVOR, ESPERE...")
    # Se realiza una pausa de diez segundos para asegurar que el proceso se complete correctamente y minimizar el riesgo de errores.
    carga_completada(10)

    # Se llama a la función `copiar_datos_a_excel`, pasándole el nombre del cine y el libro de Excel para que los datos sean copiados a la hoja de cálculo.
    copiar_datos_a_excel(nombre_cine, libro_excel)

    






#13 #Cinépolis Plaza Américas

    # Crea una instancia de WebDriverWait con un tiempo de espera máximo de 10 segundos.
    espera = WebDriverWait(driver, 10)

    # Llamada a la función elegir_ciudad_cine para seleccionar una ciudad y un cine específicos
    # valor_ciudad = "177" corresponde al atributo 'value' de una opción en el menú desplegable de ciudades
    # en el HTML, se verá algo como esto:
    # <select id="ciudad">
    #   <option value="177">Ciudad Ejemplo</option>
    #   <!-- Otras opciones -->
    # </select>
    #
    # valor_cine = "494" corresponde al atributo 'value' de una opción en el menú desplegable de cines
    # en el HTML, se verá algo como esto:
    # <select id="cine">
    #   <option value="494">Cine Ejemplo</option>
    #   <!-- Otras opciones -->
    # </select>

    nombre_cine = elegir_ciudad_cine(driver, espera, "177", "494")


    # Se imprime el nombre del cine seleccionado para confirmación.
    print(f"\n\n\n\n\n\nNOMBRE DEL CINE SELECCIONADO (13/17): {nombre_cine}.")

    # Informamos al usuario que el proceso de copia de datos desde la página web está en curso.
    print("ESTAMOS COPIANDO LOS DATOS DE LA PÁGINA WEB. POR FAVOR, ESPERE...")
    # Se realiza una pausa de diez segundos para asegurar que el proceso se complete correctamente y minimizar el riesgo de errores.
    carga_completada(10)

    # Se llama a la función `copiar_datos_a_excel`, pasándole el nombre del cine y el libro de Excel para que los datos sean copiados a la hoja de cálculo.
    copiar_datos_a_excel(nombre_cine, libro_excel)

    





#14 #Cinépolis Interplaza Xela

    # Crea una instancia de WebDriverWait con un tiempo de espera máximo de 10 segundos.
    espera = WebDriverWait(driver, 10)

    # Llamada a la función elegir_ciudad_cine para seleccionar una ciudad y un cine específicos
    # valor_ciudad = "184" corresponde al atributo 'value' de una opción en el menú desplegable de ciudades
    # en el HTML, se verá algo como esto:
    # <select id="ciudad">
    #   <option value="184">Ciudad Ejemplo</option>
    #   <!-- Otras opciones -->
    # </select>
    #
    # valor_cine = "609" corresponde al atributo 'value' de una opción en el menú desplegable de cines
    # en el HTML, se verá algo como esto:
    # <select id="cine">
    #   <option value="609">Cine Ejemplo</option>
    #   <!-- Otras opciones -->
    # </select>

    nombre_cine = elegir_ciudad_cine(driver, espera, "184", "609")


    # Se imprime el nombre del cine seleccionado para confirmación.
    print(f"\n\n\n\n\n\nNOMBRE DEL CINE SELECCIONADO (14/17): {nombre_cine}.")

    # Informamos al usuario que el proceso de copia de datos desde la página web está en curso.
    print("ESTAMOS COPIANDO LOS DATOS DE LA PÁGINA WEB. POR FAVOR, ESPERE...")
    # Se realiza una pausa de diez segundos para asegurar que el proceso se complete correctamente y minimizar el riesgo de errores.
    carga_completada(10)

    # Se llama a la función `copiar_datos_a_excel`, pasándole el nombre del cine y el libro de Excel para que los datos sean copiados a la hoja de cálculo.
    copiar_datos_a_excel(nombre_cine, libro_excel)

    









#15 #Cinépolis Utz Ulew Mall

    # Crea una instancia de WebDriverWait con un tiempo de espera máximo de 10 segundos.
    espera = WebDriverWait(driver, 10)

    # Llamada a la función elegir_ciudad_cine para seleccionar una ciudad y un cine específicos
    # valor_ciudad = "184" corresponde al atributo 'value' de una opción en el menú desplegable de ciudades
    # en el HTML, se verá algo como esto:
    # <select id="ciudad">
    #   <option value="184">Ciudad Ejemplo</option>
    #   <!-- Otras opciones -->
    # </select>
    #
    # valor_cine = "571" corresponde al atributo 'value' de una opción en el menú desplegable de cines
    # en el HTML, se verá algo como esto:
    # <select id="cine">
    #   <option value="571">Cine Ejemplo</option>
    #   <!-- Otras opciones -->
    # </select>

    nombre_cine = elegir_ciudad_cine(driver, espera, "184", "571")
    

    # Se imprime el nombre del cine seleccionado para confirmación.
    print(f"\n\n\n\n\n\nNOMBRE DEL CINE SELECCIONADO (15/17): {nombre_cine}.")

    # Informamos al usuario que el proceso de copia de datos desde la página web está en curso.
    print("ESTAMOS COPIANDO LOS DATOS DE LA PÁGINA WEB. POR FAVOR, ESPERE...")
    # Se realiza una pausa de diez segundos para asegurar que el proceso se complete correctamente y minimizar el riesgo de errores.
    carga_completada(10)

    # Se llama a la función `copiar_datos_a_excel`, pasándole el nombre del cine y el libro de Excel para que los datos sean copiados a la hoja de cálculo.
    copiar_datos_a_excel(nombre_cine, libro_excel)

    




#16 #Cinépolis Centro Gran Carchá

    # Crea una instancia de WebDriverWait con un tiempo de espera máximo de 10 segundos.
    espera = WebDriverWait(driver, 10)

    # Llamada a la función elegir_ciudad_cine para seleccionar una ciudad y un cine específicos
    # valor_ciudad = "211" corresponde al atributo 'value' de una opción en el menú desplegable de ciudades
    # en el HTML, se verá algo como esto:
    # <select id="ciudad">
    #   <option value="211">Ciudad Ejemplo</option>
    #   <!-- Otras opciones -->
    # </select>
    #
    # valor_cine = "664" corresponde al atributo 'value' de una opción en el menú desplegable de cines
    # en el HTML, se verá algo como esto:
    # <select id="cine">
    #   <option value="664">Cine Ejemplo</option>
    #   <!-- Otras opciones -->
    # </select>

    nombre_cine = elegir_ciudad_cine(driver, espera, "211", "664")


    # Se imprime el nombre del cine seleccionado para confirmación.
    print(f"\n\n\n\n\n\nNOMBRE DEL CINE SELECCIONADO (16/17): {nombre_cine}.")

    # Informamos al usuario que el proceso de copia de datos desde la página web está en curso.
    print("ESTAMOS COPIANDO LOS DATOS DE LA PÁGINA WEB. POR FAVOR, ESPERE...")
    # Se realiza una pausa de diez segundos para asegurar que el proceso se complete correctamente y minimizar el riesgo de errores.
    carga_completada(10)

    # Se llama a la función `copiar_datos_a_excel`, pasándole el nombre del cine y el libro de Excel para que los datos sean copiados a la hoja de cálculo.
    copiar_datos_a_excel(nombre_cine, libro_excel)

    




#17 #Cinépolis Pradera Zacapa

    # Crea una instancia de WebDriverWait con un tiempo de espera máximo de 10 segundos.
    espera = WebDriverWait(driver, 10)

    # Llamada a la función elegir_ciudad_cine para seleccionar una ciudad y un cine específicos
    # valor_ciudad = "232" corresponde al atributo 'value' de una opción en el menú desplegable de ciudades
    # en el HTML, se verá algo como esto:
    # <select id="ciudad">
    #   <option value="232">Ciudad Ejemplo</option>
    #   <!-- Otras opciones -->
    # </select>
    #
    # valor_cine = "719" corresponde al atributo 'value' de una opción en el menú desplegable de cines
    # en el HTML, se verá algo como esto:
    # <select id="cine">
    #   <option value="719">Cine Ejemplo</option>
    #   <!-- Otras opciones -->
    # </select>

    nombre_cine = elegir_ciudad_cine(driver, espera, "232", "719")


    # Se imprime el nombre del cine seleccionado para confirmación.
    print(f"\n\n\n\n\n\nNOMBRE DEL CINE SELECCIONADO (17/17): {nombre_cine}.")

    # Informamos al usuario que el proceso de copia de datos desde la página web está en curso.
    print("ESTAMOS COPIANDO LOS DATOS DE LA PÁGINA WEB. POR FAVOR, ESPERE...")
    # Se realiza una pausa de diez segundos para asegurar que el proceso se complete correctamente y minimizar el riesgo de errores.
    carga_completada(10)

    # Se llama a la función `copiar_datos_a_excel`, pasándole el nombre del cine y el libro de Excel para que los datos sean copiados a la hoja de cálculo.
    copiar_datos_a_excel(nombre_cine, libro_excel)

    


# Esta línea de código maneja errores que puedan ocurrir durante la ejecución del programa.
# Si ocurre algún problema (como un error inesperado), se captura el error y se muestra un mensaje.
except Exception as e:
    # Imprime un mensaje en la pantalla que indica que ocurrió un error durante la ejecución.
    # El mensaje incluye detalles sobre el error, que se almacena en la variable 'e'.
    print(f"Error en la ejecución: {e}")



# El bloque 'finally' se ejecuta siempre, sin importar si hubo un error o no.
# Esto significa que, sin importar si el código anterior terminó correctamente o si hubo un error,
# el código dentro del bloque 'finally' siempre se ejecutará al final.
finally:
    # Esta línea crea el nombre del archivo Excel que se guardará.
    # El nombre del archivo tendrá el formato "CINEPOLIS-GUATEMALA-YYYY-MM-DD.xlsx", donde YYYY-MM-DD es la fecha actual.
    # Por ejemplo, si hoy es el 12 de agosto de 2024, el nombre del archivo será "CINEPOLIS-GUATEMALA-2024-08-12.xlsx".
    nombre_archivo = f"CINEPOLIS-GUATEMALA-{datetime.now().strftime('%Y-%m-%d')}.xlsx"

    # Esta línea determina la ubicación completa del archivo que se va a guardar.
    # 'os.getcwd()' obtiene la ruta del directorio actual donde se está ejecutando el programa.
    # 'os.path.join()' combina esta ruta con el nombre del archivo para crear la ruta completa del archivo.
    # Así, 'ruta_archivo' es la dirección completa donde el archivo Excel se guardará en el sistema.
    ruta_archivo = os.path.join(os.getcwd(), nombre_archivo)


    # Muestra un mensaje en la pantalla indicando que se está guardando la información en múltiples hojas del archivo Excel.
    print("\n\n\n\nGUARDANDO LA INFORMACIÓN EN MULTIPLES HOJAS DEL ARCHIVO EXCEL...")

    # Llama a una función para eliminar iconos o símbolos en la columna H del archivo Excel.
    eliminar_icono_de_columna_h(libro_excel)

    # Hace que el programa espere 2 segundos para asegurarse de que los cambios en la columna H se hayan aplicado correctamente.
    time.sleep(2)

    # Llama a una función para eliminar palabras repetidas en la columna H del archivo Excel.
    eliminar_palabras_repetidas_columna_h(libro_excel)

    # Hace que el programa espere 2 segundos para asegurarse de que los cambios se hayan aplicado antes de continuar.
    time.sleep(2)

    # Llama a una función para convertir todo el texto en la columna E del archivo Excel a mayúsculas.
    convertir_a_mayusculas_columna_e(libro_excel)

    # Hace que el programa espere 2 segundos para asegurarse de que los cambios se hayan aplicado antes de continuar.
    time.sleep(2)

    # Llama a una función para convertir todo el texto en la columna H del archivo Excel a mayúsculas.
    convertir_a_mayusculas_columna_h(libro_excel)

    # Se realiza una pausa de cinco segundos para asegurar que el proceso se complete correctamente y minimizar el riesgo de errores.
    carga_completada(5)

    # Guarda el archivo de Excel con todos los datos y cambios que hemos hecho.
    # 'ruta_archivo' es la ubicación y nombre del archivo donde se guardará.
    libro_excel.save(ruta_archivo)

    # Imprime un mensaje en la consola que indica que el archivo de Excel se ha guardado correctamente.
    # 'nombre_archivo' muestra el nombre del archivo donde se guardaron las hojas.
    print(f"¡LAS HOJAS EXCEL SE HAN GUARDADO EN EL ARCHIVO EXCEL: {nombre_archivo}.!")

    # Pausa la ejecución del programa y espera a que el usuario presione ENTER antes de continuar.
    # Esto permite al usuario leer el mensaje anterior antes de cerrar el navegador.
    input("\n\n Presiona ENTER para cerrar el navegador y salir de la consola Python.")

    # Cierra el navegador web y termina la sesión del controlador (driver).
    # Esto es necesario para liberar los recursos del sistema una vez que hemos terminado.
    driver.quit()

